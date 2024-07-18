# Importación de módulos para el manejo de imágenes
from PIL import Image, ImageTk, ImageWin, ImageDraw  # Para trabajar con imágenes en Tkinter

# Importación del módulo tkinter para la creación de la interfaz gráfica
import tkinter as tk
from tkinter import ttk
from tkinter import Scrollbar  # Importación específica de la barra de desplazamiento
from tkinter import messagebox  # Para mostrar mensajes de cuadro de diálogo
from tkinter import filedialog
from tkinter import simpledialog

# Importación de módulos para el manejo de archivos Excel
from openpyxl import load_workbook, Workbook

# Importación de módulos para la generación de archivos PDF
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

# Importación del widget de calendario para Tkinter
from tkcalendar import DateEntry
#Importancion de pandas para manipulacion de info de excel
import pandas as pd
# Importación del módulo 'os' para operaciones del sistema
import os
#importacion para graficar
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('TkAgg')


# Importar el módulo fitz de PyMuPDF
import fitz
import json


#Importar metodos para mandar correo
import smtplib
import email.mime.multipart
import email.mime.base
import email.mime.text
#clase para dibujar la firma
class SignatureApp:
    def __init__(self, root, save_name):
        self.root = root
        self.save_name = save_name
        self.root.title(f"Firma con el Mouse - {save_name}")

        self.canvas = tk.Canvas(self.root, width=400, height=200, bg="white")
        self.canvas.pack()

        self.button_frame = tk.Frame(self.root)
        self.button_frame.pack(fill=tk.X)

        self.clear_button = tk.Button(self.button_frame, text="Limpiar", command=self.clear_canvas)
        self.clear_button.pack(side=tk.LEFT)

        self.save_button = tk.Button(self.button_frame, text="Guardar", command=self.save_signature)
        self.save_button.pack(side=tk.LEFT)

        self.canvas.bind("<B1-Motion>", self.paint)

        self.image = Image.new("RGB", (400, 200), "white")
        self.draw = ImageDraw.Draw(self.image)
    #Dibujar
    def paint(self, event):
        x1, y1 = (event.x - 1), (event.y - 1)
        x2, y2 = (event.x + 1), (event.y + 1)
        self.canvas.create_oval(x1, y1, x2, y2, fill="black", width=2)
        self.draw.line([x1, y1, x2, y2], fill="black", width=2)
        #limpiar lienzo
    def clear_canvas(self):
        self.canvas.delete("all")
        self.image = Image.new("RGB", (400, 200), "white")
        self.draw = ImageDraw.Draw(self.image)
        #Guardar Firma
    def save_signature(self):
        filename = f"firma_{self.save_name}.png"
        self.image.save(filename)
        messagebox.showinfo("Firma", "Firma guardada correctamente.")
        self.root.destroy()  # Cierra la ventana de firma después del mensaje

class BitacoraMantenimiento:
    def __init__(self, root):   
        
        self.root = root
        self.root.title("Bitácora de Mantenimiento")
        
       
        # Bloquear la opción de maximizar la ventana principal
        self.root.resizable(False, False)
      
        # Crear un lienzo para colocar el frame con la barra de desplazamiento
        canvas = tk.Canvas(root)
        canvas.pack(side='left', fill='both', expand=True)
        
        # Crear un frame dentro del lienzo
        self.frame = tk.Frame(canvas)
        self.frame.pack(padx=10, pady=10)
        
        # Añadir la barra de desplazamiento vertical
        scrollbar = tk.Scrollbar(root, orient='vertical', command=canvas.yview)
        scrollbar.pack(side='right', fill='y')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Configurar el lienzo para usar la barra de desplazamiento
        canvas.create_window((0, 0), window=self.frame, anchor='nw')
        
        # Vincular la barra de desplazamiento con el movimiento del ratón
        self.frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        # Enlazar el desplazamiento del mouse al lienzo
        canvas.bind_all("<MouseWheel>", lambda event: canvas.yview_scroll(int(-1*(event.delta/120)), "units"))

        # Agregar contenido al frame
        self.add_content()
        

    def add_content(self):
        
        
        # Intentar cargar el número guardado desde el archivo
        self.numero_guardado = self.cargar_numero()
        #self.generar_codigo_barras(str(self.numero_guardado))
        # Crear una etiqueta para mostrar el contador
        self.contador_label = tk.Label(root, text=f"{self.numero_guardado}")
        self.contador_label.place(x=850, y=10)  # Colocar la etiqueta en la esquina superior derecha
        
                  # Cargar las imágenes
        self.logo1_image = Image.open("./img/logo1.png")
        self.logo1_image = self.logo1_image.resize((180, 100), )  # Redimensionar la imagen
        self.logo1_photo = ImageTk.PhotoImage(self.logo1_image)
        
        self.logo2_image = Image.open("./img/logo2.png")
        self.logo2_image = self.logo2_image.resize((180, 100), )  # Redimensionar la imagen
        self.logo2_photo = ImageTk.PhotoImage(self.logo2_image)
        
        # Etiqueta para la primera imagen
        self.logo1_label = tk.Label(self.frame, image=self.logo1_photo)
        self.logo1_label.grid(row=0, column=0, padx=10, pady=10)
        
        # Etiqueta para la segunda imagen
        self.logo2_label = tk.Label(self.frame, image=self.logo2_photo)
        self.logo2_label.grid(row=0, column=3, padx=10, pady=10)

        # Título para la sección de Información del Equipo
        titulo_informacion_equipo = tk.Label(self.frame, text="Bitacora de Mantenimiento", font=("Arial", 14, "bold"), pady=10)
        titulo_informacion_equipo.grid(row=0, column=0, columnspan=4)
        titulo_informacion_equipo.config(justify="center")
        
        # Campos de entrada
        tk.Label(self.frame, text="Fecha de Recepción:").grid(row=1, column=0, sticky="e")
        self.fecha_recepcion_entry = DateEntry(self.frame, width=12, background='darkblue', foreground='white', borderwidth=2)
        self.fecha_recepcion_entry.grid(row=1, column=1)
        
        tk.Label(self.frame, text="Fecha de Entrega:").grid(row=1, column=2, sticky="e")
        self.fecha_entrega_entry = DateEntry(self.frame, width=12, background='darkblue', foreground='white', borderwidth=2)
        self.fecha_entrega_entry.grid(row=1, column=3)
        
        # Título 
        titulo_datos_equipo = tk.Label(self.frame, text="Datos del Responsable del Equipo", font=("Helvetica", 12, "bold"), pady=10)
        titulo_datos_equipo.grid(row=2, column=0, columnspan=4)
        titulo_datos_equipo.config(justify="center")
        
        tk.Label(self.frame, text="Nombre:").grid(row=3, column=0, sticky="e")
        self.nombre_responsable_entry = tk.Entry(self.frame)
        self.nombre_responsable_entry.grid(row=3, column=1)
        
        tk.Label(self.frame, text="Teléfono:").grid(row=4, column=0, sticky="e")
        self.telefono_responsable_entry = tk.Entry(self.frame)
        self.telefono_responsable_entry.grid(row=4, column=1)
        
        
        
        tk.Label(self.frame, text="Correo:").grid(row=3, column=2, sticky="e")
        self.correo_entry = tk.Entry(self.frame)
        self.correo_entry.grid(row=3, column=3)

                
        tk.Label(self.frame, text="Área:").grid(row=4, column=2, sticky="e")
        self.area_responsable_entry = tk.Entry(self.frame)
        self.area_responsable_entry.grid(row=4, column=3)
        
        # Título 
        titulo_descripcion_equipo = tk.Label(self.frame, text="Descripcion de Equipo", font=("Helvetica", 12, "bold"), pady=10)
        titulo_descripcion_equipo.grid(row=6, column=0, columnspan=4)
        titulo_descripcion_equipo.config(justify="center")
        # Crear el Label "Equipo:"
        tk.Label(self.frame, text="Equipo:").grid(row=7, column=0, sticky="e")
        
        # Definir opciones para el Combobox
        equipo_options = ["Autoclave", "Balanza digital", "Balanza Granataria", "Basculas", "Bocina", "Bomba de agua", "Bomba de vacío", 
                          "Campana", "Cámaras", "Cargador", "Centrifugadora", "Cortadora", "Compresora", "Computadora", "Control Remoto", "Cronometro", 
                          "Dispensador de agua", "Diseño 3D", "Eliminador", "Esmeril", "Espectrofotómetro", "Estufa", "Fuentes de poder", "Generador de funciones", 
                          "Horno", "Impresión 3D", "Impresora", "Incubadora", "Lámpara", "Licuadora", "Lavadoras", "Mano factura de piezas", "Manta de calentamiento", 
                          "Microscopio", "Microscopio Estereoscópico", "Monitor", "Motor", "Mufla", "Multímetro", "Multi-contactos", "No break", "Osciloscopio","Pantalla", 
                          "Parilla con agitación", "Parrilla de calentamiento", "Parrilla de calentamiento y agitación", "Pipetas", "Proyector", "Pulidora", "Purificador", 
                          "Radio", "Refrigerador", "Regulador de Voltaje", "Sensores", "Termómetro digital", "Taladro", "Torneo y fresado", "Tornos","Torniquete","UPS","Otros" ]
        
        # Crear el Combobox y posicionarlo a la derecha del Label
        self.equipo_combobox = ttk.Combobox(self.frame, values=equipo_options)
        self.equipo_combobox.grid(row=7, column=1, sticky="w")
        self.equipo_combobox.bind("<<ComboboxSelected>>", self.toggle_equipo)
        
        tk.Label(self.frame, text="Otros:").grid(row=7, column=2, sticky="e")
        self.otro_equipo_entry = tk.Entry(self.frame, state="disabled")
        self.otro_equipo_entry.grid(row=7, column=3)
        
        tk.Label(self.frame, text="Modelo:").grid(row=8, column=0, sticky="e")
        self.modelo_equipo_entry = tk.Entry(self.frame)
        self.modelo_equipo_entry.grid(row=8, column=1)
        
        tk.Label(self.frame, text="Marca:").grid(row=8, column=2, sticky="e")
        self.marca_equipo_entry = tk.Entry(self.frame)
        self.marca_equipo_entry.grid(row=8, column=3)
        
        tk.Label(self.frame, text="No. Serie").grid(row=10, column=0, sticky="e")
        self.no_serie_var = tk.StringVar(value="N/A")
        self.no_serie_entry = tk.Entry(self.frame, textvariable=self.no_serie_var)
        self.no_serie_entry.grid(row=10, column=1)
        
        tk.Label(self.frame, text="No. Inventario").grid(row=10, column=2, sticky="e")
        self.no_inventario_var = tk.StringVar(value="N/A")
        self.no_inventario_entry = tk.Entry(self.frame, textvariable=self.no_inventario_var)
        self.no_inventario_entry.grid(row=10, column=3)
        
        # Título 
        titulo_estado_equipo = tk.Label(self.frame, text="Estado del Equipo", font=("Helvetica", 12, "bold"), pady=10)
        titulo_estado_equipo.grid(row=11, column=0, columnspan=4)
        titulo_estado_equipo.config(justify="center")
        
        # Sección de Estado del Equipo
        tk.Label(self.frame, text="Estado del Equipo:").grid(row=12, column=0, sticky="e")
        self.estado_enciende = tk.BooleanVar()
        self.enciende_checkbutton = tk.Checkbutton(self.frame, text="Enciende", variable=self.estado_enciende)
        self.enciende_checkbutton.grid(row=12, column=1, sticky="w")
        
        self.estado_cable = tk.BooleanVar()
        self.cable_checkbutton = tk.Checkbutton(self.frame, text="Cable de Alimentación", variable=self.estado_cable)
        self.cable_checkbutton.grid(row=12, column=2, sticky="w")
        
        self.estado_componente = tk.BooleanVar()
        self.componente_checkbutton = tk.Checkbutton(self.frame, text="Falta algún componente", variable=self.estado_componente,command=self.toggle_falta)
        self.componente_checkbutton.grid(row=13, column=1, sticky="w")
        
        
        self.faltante_equipo_entry = tk.Entry(self.frame, state="disabled")
        self.faltante_equipo_entry.grid(row=13, column=3)
        
        
        self.estado_dano_botones = tk.BooleanVar()
        self.dano_botones_checkbutton = tk.Checkbutton(self.frame, text="Daño en botones/perillas", variable=self.estado_dano_botones)
        self.dano_botones_checkbutton.grid(row=13, column=2, sticky="w")
        
        self.estado_corrosion = tk.BooleanVar()
        self.corrosion_checkbutton = tk.Checkbutton(self.frame, text="Presenta corrosión/oxidación", variable=self.estado_corrosion)
        self.corrosion_checkbutton.grid(row=14, column=1, sticky="w")
        
        self.estado_dano_carcasa = tk.BooleanVar()
        self.dano_carcasa_checkbutton = tk.Checkbutton(self.frame, text="Daño en Carcasa", variable=self.estado_dano_carcasa)
        self.dano_carcasa_checkbutton.grid(row=14, column=2, sticky="w")
        
        # Título 
        titulo_estado_equipo = tk.Label(self.frame, text="Descripción Detallada del Equipo", font=("Helvetica", 12, "bold"), pady=10)
        titulo_estado_equipo.grid(row=15, column=0, columnspan=4)
        titulo_estado_equipo.config(justify="center")
        
        # Otros campos de información
        tk.Label(self.frame, text="Descripción Detallada del Equipo:").grid(row=16, column=0, sticky="ne")
        self.descripcion_detallada_entry = tk.Text(self.frame, height=5, width=80, wrap=tk.WORD)
        self.descripcion_detallada_entry.grid(row=16, column=1, columnspan=3, sticky="w")
        
        # Título 
        titulo_estado_equipo = tk.Label(self.frame, text="Mantenimineto", font=("Helvetica", 12, "bold"), pady=10)
        titulo_estado_equipo.grid(row=17, column=0, columnspan=4)
        titulo_estado_equipo.config(justify="center")
        
        # Sección de Mantenimiento del Equipo
        tk.Label(self.frame, text="Mantenimiento del Equipo:").grid(row=18, column=0, sticky="e")
        self.preventivo_var = tk.BooleanVar()
        self.preventivo_checkbutton = tk.Checkbutton(self.frame, text="Preventivo", variable=self.preventivo_var)
        self.preventivo_checkbutton.grid(row=18, column=1, sticky="w")
        
        self.correctivo_var = tk.BooleanVar()
        self.correctivo_checkbutton = tk.Checkbutton(self.frame, text="Correctivo", variable=self.correctivo_var)
        self.correctivo_checkbutton.grid(row=18, column=2, sticky="w")
        
        self.diagnostico_var = tk.BooleanVar()
        self.diagnostico_checkbutton = tk.Checkbutton(self.frame, text="Diagnóstico", variable=self.diagnostico_var)
        self.diagnostico_checkbutton.grid(row=19, column=1, sticky="w")
        
        self.puesta_en_marcha_var = tk.BooleanVar()
        self.puesta_en_marcha_checkbutton = tk.Checkbutton(self.frame, text="Puesta en Marcha", variable=self.puesta_en_marcha_var)
        self.puesta_en_marcha_checkbutton.grid(row=19, column=2, sticky="w")
        
        self.otro_var = tk.BooleanVar()
        self.otro_checkbutton = tk.Checkbutton(self.frame, text="Otro", variable=self.otro_var)
        self.otro_checkbutton.grid(row=19, column=3, sticky="w")
        
        # Otros campos de información
        tk.Label(self.frame, text="Descripción:").grid(row=20, column=0, sticky="ne")
        self.descripcion_entry = tk.Text(self.frame, height=5, width=80, wrap=tk.WORD)
        self.descripcion_entry.grid(row=20, column=1, columnspan=3, sticky="w")
        
        tk.Label(self.frame, text="¿Fue reparado?").grid(row=21, column=0, sticky="e")
        self.reparado_var = tk.StringVar()
        reparado_options = ["Si", "No"]
        self.reparado_dropdown = tk.OptionMenu(self.frame, self.reparado_var, *reparado_options)
        self.reparado_dropdown.grid(row=21, column=1, columnspan=2, sticky="w")
        
        
        
        # Título 
        titulo_estado_equipo = tk.Label(self.frame, text="Lista de Materiales Utilizados", font=("Helvetica", 12, "bold"), pady=10)
        titulo_estado_equipo.grid(row=22, column=0, columnspan=4)
        titulo_estado_equipo.config(justify="center")
        
        # Checkboxes para Materiales Utilizados
        tk.Label(self.frame, text="Materiales Utilizados:").grid(row=23, column=0, sticky="e")
        self.materiales_utilizados = [
            ("Aceite lubricante multiusos", tk.BooleanVar()),
            ("Alcohol Isopropílico", tk.BooleanVar()),
            ("Soldadura (Estaño)", tk.BooleanVar()),
            ("Aislantes", tk.BooleanVar()),
            ("Cable de Conexión", tk.BooleanVar()),
            ("Conectores y/o terminales", tk.BooleanVar()),
            ("Potenciómetros", tk.BooleanVar()),
            ("Cables de Alimentación", tk.BooleanVar()),
            ("Dispositivos electrónicos de potencia", tk.BooleanVar()),
            ("Dispositivos de sujeción", tk.BooleanVar()),
            ("Fusibles", tk.BooleanVar()),
            ("Liquido Limpiador Multiusos", tk.BooleanVar())
        ]
        for i, (nombre, var) in enumerate(self.materiales_utilizados):
            if i < 6:
                column = 1
            else:
                column = 2
                i -= 6
            tk.Checkbutton(self.frame, text=nombre, variable=var).grid(row=23+i, column=column, sticky="w")
            
        self.otro_var = tk.BooleanVar()
        self.otro_checkbox = tk.Checkbutton(self.frame, text="Otros", variable=self.otro_var,command=self.toggle_otros)
        self.otro_checkbox.grid(row=30, column=2, sticky="w")   
        
        # Caja de texto para Otros Materiales Utilizados
        tk.Label(self.frame, text="Otros:").grid(row=30, column=0, sticky="e")
        self.otros_materiales_entry = tk.Entry(self.frame, state="disabled")
        self.otros_materiales_entry.grid(row=30, column=1, columnspan=2, sticky="w")
    
        # Título 
        titulo_estado_equipo = tk.Label(self.frame, text="Responsables", font=("Helvetica", 12, "bold"), pady=10)
        titulo_estado_equipo.grid(row=31, column=0, columnspan=4)
        titulo_estado_equipo.config(justify="center")
          
                # Responsable de Taller
        tk.Label(self.frame, text="Responsable de Taller:").grid(row=32, column=0, sticky="e")
        self.responsable_taller_var = tk.StringVar()
        responsable_taller_options = ["Juan Daniel Ramírez Zamora", "Jose Manuel Fernandez Ramírez", "Otro"]
        self.responsable_taller_dropdown = tk.OptionMenu(self.frame, self.responsable_taller_var, *responsable_taller_options)
        self.responsable_taller_dropdown.grid(row=32, column=1, columnspan=2, sticky="w")
        
        
        
        # Cajas de texto para Responsable equipo Recepción
        
        t2=tk.Label(self.frame, text="¿Recoge la misma persona?").grid(row=32, column=2, sticky="w")
        self.recep_var = tk.StringVar()
        recep_options = ["Si", "No"]
        self.recep_dropdown = tk.OptionMenu(self.frame, self.recep_var, *recep_options,command=self.toggle_recibir)
        self.recep_dropdown.grid(row=33, column=2, columnspan=2, sticky="w")
        
        
        tk.Label(self.frame, text="Responsable equipo Recepción:").grid(row=33, column=0, sticky="e")
        self.responsable_recepcion_entry = tk.Entry(self.frame)
        self.responsable_recepcion_entry.grid(row=33, column=1, columnspan=2, sticky="w")
        self.responsable_recepcion_entry.config(state="disabled")
        

       
        # Definir la fila para los botones
        fila_botones = 34
        
        # Ajustar el ancho de cada columna
        self.frame.grid_columnconfigure((0, 1, 2, 3, 4, 5, 6), weight=1, uniform="equal")
        
        # Botón para generar gráficos
        #self.graficos_button = tk.Button(self.frame, text="Generar Gráficas", command=self.generar_graficas)
        self.estadisticas_button = tk.Button(self.frame, text="Estadísticas", command=self.abrir_ventana_estadisticas)
        self.estadisticas_button.grid(row=fila_botones, column=0, padx=3, pady=10, sticky="ew")
        
        # Botón para guardar
        self.guardar_button = tk.Button(self.frame, text="Guardar en Excel", command=lambda: self.guardar_bitacora())
        self.guardar_button.grid(row=fila_botones, column=1, padx=3, pady=10, sticky="ew")
        
        # Botón para guardar progreso
        self.guardar_p_button = tk.Button(self.frame, text="Guardar Progreso", command=lambda: self.guardar_datos())
        self.guardar_p_button.grid(row=fila_botones, column=2, padx=3, pady=10, sticky="ew")

               
        # Botón para generar ticket
        self.generar_ticket_button = tk.Button(self.frame, text="Generar Ticket", command=lambda: self.generar_ticket())
        self.generar_ticket_button.grid(row=35, column=1, padx=3, pady=10, sticky="ew")
        # Botón para generar ticket
        self.correo_button = tk.Button(self.frame, text="Mandar Correo", command=self.mandar_correo)
        self.correo_button.grid(row=35, column=2, padx=3, pady=10, sticky="ew")
        #Boton abrir vetana firma A
        self.firma_taller = tk.Button(self.frame, text="Firma Responsable Taller", command=lambda: self.open_signature_app("Taller"))
        self.firma_taller.grid(row=34, column=3, padx=4, pady=10, sticky="ew")
        #boton abririr ventan firma B
        self.firma_equipo = tk.Button(self.frame, text="Firma Responsable Equipo", command=lambda: self.open_signature_app("Equipo"))
        self.firma_equipo.grid(row=35, column=3, padx=4, pady=10, sticky="ew")

        # Variable de control para el botón de ticket
        self.ticket_presionado = False
        self.correo_presionado= False
        
        # Desactivar el botón de guardar al principio
        self.guardar_button.config(state="disabled")
        # Desactivar el botón de guardar al principio
        self.generar_ticket_button.config(state="disabled")
        self.correo_button.config(state="disabled")
        self.firma_taller.config(state="disabled")
        self.estadisticas_button.config(state="disabled")
        self.guardar_p_button.config(state="disabled")
        self.cargar_datos()
    
    #Metodo para abrir el cuadro de firma
    def open_signature_app(self, name):
        signature_window = tk.Toplevel(self.root)
        app = SignatureApp(signature_window, name)
        self.firma_taller.config(state="normal")
        self.firma_equipo.config(state="disabled")
        if name=="Taller":
            self.firma_taller.config(state="disable")
            self.generar_ticket_button.config(state="normal")
            
        
    def generar_graficas(self):
        # Paso 1: Leer el archivo Excel
        df = pd.read_excel('bitacora_mantenimiento.xlsx')
    
        # Paso 2: Separar los datos por comas en una nueva columna para los tipos de mantenimiento
        df['Tipos de mantenimiento separados'] = df['Tipos de Mantenimiento'].str.split(',')
        df = df.explode('Tipos de mantenimiento separados')
        frecuencia_tipos_mantenimiento = df['Tipos de mantenimiento separados'].value_counts()
    
        # Paso 3: Crear el gráfico de barras para tipos de mantenimiento
        plt.figure(figsize=(10, 5))  
        frecuencia_tipos_mantenimiento.plot(kind='bar')
        # Paso 4: Rotar los labels en el eje x
        plt.xticks(rotation=45)
        plt.xlabel('Tipos de mantenimiento')
        plt.ylabel('Frecuencia')
        plt.title('Frecuencia de Tipos de mantenimiento')
        plt.show()
    
        # Paso 4: Crear el gráfico de pastel para la columna "¿Fue Reparado?"
        plt.figure(figsize=(8, 8))  
        reparado_frecuencia = df['¿Fue Reparado?'].value_counts()
        reparado_frecuencia.plot(kind='pie', autopct='%1.1f%%')
        plt.title('Distribución de ¿Fue Reparado?')
        plt.show()
    
        # Paso 5: Separar los datos de materiales utilizados por comas en una nueva columna
        df['Materiales utilizados separados'] = df['Lista de Materiales Utilizados'].str.split(',')
        df = df.explode('Materiales utilizados separados')
        frecuencia_materiales = df['Materiales utilizados separados'].value_counts()
    
        # Paso 6: Crear el gráfico de barras para materiales utilizados
        plt.figure(figsize=(10, 5))  
        frecuencia_materiales.plot(kind='bar')
        # Paso 4: Rotar los labels en el eje x
        plt.xticks(rotation=45)
        plt.xlabel('Materiales utilizados')
        plt.ylabel('Frecuencia')
        plt.title('Frecuencia de Materiales Utilizados')
        plt.show()
    
        # Paso 7: Crear el gráfico de barras para la columna Equipo
        plt.figure(figsize=(10, 5))
        equipo_frecuencia = df['Equipo'].value_counts()
        equipo_frecuencia.plot(kind='bar')
        plt.xticks(rotation=45)
        plt.xlabel('Equipo')
        plt.ylabel('Frecuencia')
        plt.title('Frecuencia de Equipos')
        plt.show()
        
    def abrir_ventana_estadisticas(self):
        # Crear una nueva ventana para las estadísticas
        self.ventana_estadisticas = tk.Toplevel(self.root)
        self.ventana_estadisticas.title("Estadísticas")

        # Definir los botones en la ventana de estadísticas
        self.crear_botones_estadisticas()

    def crear_botones_estadisticas(self):
        # Botón para generar gráficas
        self.graficas_button = tk.Button(self.ventana_estadisticas, text="Generar Gráficas", command=self.generar_graficas)
        self.graficas_button.pack(pady=10)

        # Botón para abrir Excel
        self.abrir_excel_button = tk.Button(self.ventana_estadisticas, text="Abrir Excel", command=self.abrir_excel)
        self.abrir_excel_button.pack(pady=10)

        # Botón para abrir PDF
        self.abrir_pdf_button = tk.Button(self.ventana_estadisticas, text="Abrir PDF", command=self.open_pdf)
        self.abrir_pdf_button.pack(pady=10)
    
    def mandar_correo(self):

        # Obtener el correo electrónico del campo de entrada
        correo_destino = self.correo_entry.get()
        
        # Configuración del servidor SMTP
        servidor_smtp = 'smtp.gmail.com'
        puerto_smtp = 587

        # Dirección de correo electrónico y contraseña del remitente
        remitente = 'uclreparaciones@gmail.com'
        password = 'lufs mesw gxyw kvzi'

        try:
            # Crea la conexión SMTP
            server = smtplib.SMTP(servidor_smtp, puerto_smtp)
            server.starttls()  # Habilitar cifrado TLS
            server.login(remitente, password)  # Iniciar sesión en el servidor SMTP

            # Definir el remitente y destinatario del correo electrónico
            
            remitente = remitente
            destinatario = correo_destino

            # Crear el mensaje del correo electrónico
            mensaje = email.mime.multipart.MIMEMultipart()
            mensaje['From'] = remitente
            mensaje['To'] = destinatario
            mensaje['Subject'] = "Correo electrónico con archivo adjunto"

            # Añadir el cuerpo del mensaje
            cuerpo = "Estimad@ Usuari@ \nAdjunto a este correo electrónico, se envía el documento que certifica el mantenimiento preventivo, correctivo o fabricación realizado en su equipo o pieza por el Taller de Reparaciones de la Unidad Central de Laboratorios.\nSaludos cordiales \nEn caso de cualquier duda o aclaración contactarse a:\nreparaciones_ucl@uaeh.edu.mx \no a la ext.: 13224"
            mensaje.attach(email.mime.text.MIMEText(cuerpo, 'plain'))

            # Añadir el archivo como adjunto
            folio= self.contador_label.cget("text")
            ruta_archivo = f'./Archivos/PDFs/{folio}.pdf'
            with open(ruta_archivo, 'rb') as archivo:
                adjunto = email.mime.base.MIMEBase('application', 'octet-stream')
                adjunto.set_payload(archivo.read())
            email.encoders.encode_base64(adjunto)
            adjunto.add_header('Content-Disposition', f"attachment; filename= {ruta_archivo}")
            mensaje.attach(adjunto)

            # Convertir el mensaje a texto plano
            texto = mensaje.as_string()

            # Enviar el correo electrónico
            server.sendmail(remitente, destinatario, texto)

            # Cerrar la conexión SMTP
            server.quit()

            print("Correo enviado exitosamente")
            messagebox.showinfo("Correo", "Correo enviado exitosamente")
            self.correo_presionado=True
            
            self.correo_button.config(state="disabled")
            self.guardar_p_button.config(state="normal")
            # Habilitar el botón de mandar correo
        except Exception as e:
            print(f"Error al enviar el correo electrónico: {e}")
            messagebox.showinfo("Correo", f"Error al enviar el correo electrónico: {e}")
            
    def toggle_equipo(self, event):
        selected_item = self.equipo_combobox.get()  # Obtener el elemento seleccionado
        if selected_item == "Otros":
            self.otro_equipo_entry.config(state="normal")  # Activar la caja de texto "otro_equipo"
        else:
            self.otro_equipo_entry.config(state="disabled")  # Desactivar la caja de texto "otro_equipo"
    


                    
    def open_pdf(self):
        # Abrir el cuadro de diálogo para seleccionar un archivo PDF
        file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    
        if not file_path:
            return
    
        # Crear una nueva ventana para mostrar el PDF
        pdf_window = tk.Toplevel(self.root)
        pdf_window.title("PDF Viewer")
        pdf_window.geometry("650x700")  # Cambia el tamaño de la ventana a 800x600 píxeles
    
        # Crear un widget Canvas para contener el frame
        canvas = tk.Canvas(pdf_window)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
        # Crear una barra de desplazamiento vertical
        scrollbar = tk.Scrollbar(pdf_window, orient=tk.VERTICAL, command=canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.config(yscrollcommand=scrollbar.set)
    
        # Crear un frame para contener las imágenes del PDF
        pdf_frame = tk.Frame(canvas)
        canvas.create_window((0, 0), window=pdf_frame, anchor=tk.NW)
    
        # Función para desplazar con la rueda del mouse
        def on_mousewheel(event):
            canvas.yview_scroll(-1 * int((event.delta / 120)), "units")
    
        # Vincular el evento de la rueda del mouse al desplazamiento
        pdf_window.bind_all("<MouseWheel>", on_mousewheel)
    
        # Abrir el PDF seleccionado
        pdf_document = fitz.open(file_path)
    
        # Mostrar cada página del PDF en el frame
        for page_number in range(pdf_document.page_count):
            page = pdf_document.load_page(page_number)
            pix = page.get_pixmap()
            width, height = pix.width, pix.height
    
            # Convertir la imagen de PIL en una imagen de tkinter
            img = Image.frombytes("RGB", [width, height], pix.samples)
            tk_img = ImageTk.PhotoImage(img)
    
            # Mostrar la imagen en el frame
            page_label = tk.Label(pdf_frame, image=tk_img)
            page_label.pack(side=tk.TOP, padx=10, pady=10)
            page_label.image = tk_img
    
        pdf_document.close()  # Cerrar el documento después de usarlo
    
        
      
    def generar_ticket(self):
        if self.reparado_var.get() == "No":  # Si no fue reparado
        # Definir la función para guardar el motivo
            def guardar_motivo(otros_motivo_entry):
                motivo = ""
                for motivo_check, motivo_var in self.motivos_no_reparado:
                    if motivo_var.get():
                        motivo += motivo_check + ", "
                otro_motivo = otros_motivo_entry.get()
                if otro_motivo:
                    motivo += otro_motivo + ", "
        
                # Obtener el número de contador
                numero = self.contador_label.cget("text")
        
                # Nombre del archivo Excel
                archivo_excel = "Bitacora_No_Reparados.xlsx"
        
                # Verificar si el archivo existe
                if not os.path.exists(archivo_excel):
                    # Crear un nuevo archivo Excel si no existe
                    libro_excel = Workbook()
                    hoja_motivos = libro_excel.active
                    hoja_motivos.title = "Motivos_No_Reparado1"
                    hoja_motivos.append(["Numero", "Motivo"])
        
                    # Guardar el numero y el motivo en la hoja
                    hoja_motivos.append([numero, motivo])
        
                    # Guardar cambios en el archivo
                    libro_excel.save(archivo_excel)
                else:
                    # Abrir el archivo Excel existente
                    libro_excel = load_workbook(archivo_excel)
                    hoja_motivos = libro_excel.active
        
                    # Encontrar la última fila en la hoja
                    last_row = hoja_motivos.max_row
        
                    # Guardar el numero y el motivo en la hoja
                    hoja_motivos.append([numero, motivo])
        
                    # Guardar cambios en el archivo
                    libro_excel.save(archivo_excel)
        
                # Mostrar mensaje de motivo guardado
                messagebox.showinfo("Motivo Guardado", f"El motivo por el cual no fue reparado es: {motivo}")
        
                # Cerrar la ventana del motivo
                motivo_window.destroy()
                selected_item = self.equipo_combobox.get()
                # Generar el ticket
                numero = self.contador_label.cget("text")
                
                ticket_pdf_path =  f"./Archivos/PDFs/{numero}.pdf"
                c = canvas.Canvas(ticket_pdf_path, pagesize=letter)
                # Establecer la fuente en el lienzo
                
                c.drawString(200, 750, "Ticket Generado")
                c.drawString(100, 730, "Número de folio: {}".format(numero))
                c.drawString(100, 710, "Fecha de Recepcion: {}".format(self.fecha_recepcion_entry.get()))
                c.drawString(100, 690, "Fecha de Entrega: {}".format(self.fecha_entrega_entry.get()))
                c.drawString(100, 670, "Nombre del Responsable: {}".format(self.nombre_responsable_entry.get()))
                c.drawString(100, 650, "Teléfono del Responsable: {}".format(self.telefono_responsable_entry.get()))
                c.drawString(100, 630, "Área del Responsable: {}".format(self.area_responsable_entry.get()))
                c.drawString(100, 610, "Equipo: {}".format(selected_item))
                c.drawString(100, 590, "Otro Equipo: {}".format(self.otro_equipo_entry.get()))
                c.drawString(100, 570, "Modelo del Equipo: {}".format(self.modelo_equipo_entry.get()))
                c.drawString(100, 550, "Marca del Equipo: {}".format(self.marca_equipo_entry.get()))
                c.drawString(100, 530, "Número de Serie: {}".format(self.no_serie_entry.get()))
                c.drawString(100, 510, "Número de Inventario: {}".format(self.no_inventario_entry.get()))
                falta_equipo= self.faltante_equipo_entry.get()
                c.drawString(200, 490, "Estado Equipo")
                c.drawString(100, 470, "Enciende: {}".format("Sí" if self.estado_enciende.get() else "No"))
                c.drawString(100, 450, "Falta algun componente: {}".format((falta_equipo) if self.estado_componente.get() else "No"))
                c.drawString(100, 430, "Presenta Corrosion/Oxidacion: {}".format("Sí" if self.estado_corrosion.get() else "No"))
                c.drawString(100, 410, "Cable Alimentacion: {}".format("Sí" if self.estado_cable.get() else "No"))
                c.drawString(100, 390, "Daño Botones/Perillas: {}".format("Sí" if self.estado_dano_botones.get() else "No"))
                c.drawString(100, 370, "Daño Carcasa: {}".format("Sí" if self.estado_dano_carcasa.get() else "No"))
                c.drawString(100, 350, "Descripcion detallada: {}".format(self.descripcion_detallada_entry.get("1.0", tk.END).strip()))
                #c.drawString(200, 330, "Mantenimineto:")
                
                # Obtener los tipos de mantenimiento seleccionados
                tipos_mantenimiento = []
                if self.preventivo_var.get():
                    tipos_mantenimiento.append("Preventivo")
                if self.correctivo_var.get():
                    tipos_mantenimiento.append("Correctivo")
                if self.diagnostico_var.get():
                    tipos_mantenimiento.append("Diagnóstico")
                if self.puesta_en_marcha_var.get():
                    tipos_mantenimiento.append("Puesta en Marcha")
                if self.otro_var.get():
                    tipos_mantenimiento.append("Otro")
                
                # Dibujar los tipos de mantenimiento en el PDF
                tipos_mantenimiento_chunks = [tipos_mantenimiento[i:i+3] for i in range(0, len(tipos_mantenimiento), 3)]
                for index, chunk in enumerate(tipos_mantenimiento_chunks):
                    c.drawString(100, 330 - index * 20, "Tipos de Mantenimiento {}: {}".format(index+1, ", ".join(chunk)))
                
                c.drawString(100, 290, "Descripcion: {}".format(self.descripcion_entry.get("1.0", tk.END).strip()))
                c.drawString(100, 270, "¿Fue Reparado?: {}".format(self.reparado_var.get()))
                c.drawString(200, 250, "Materiales Utilizados")
                
                # Obtener los materiales utilizados seleccionados
                otros_materiales = self.otros_materiales_entry.get()
                materiales_utilizados = [nombre for nombre, var in self.materiales_utilizados if var.get()]
                materiales_utilizados_completos = materiales_utilizados.copy()  # Copia la lista original
                if otros_materiales:
                    materiales_utilizados_completos.append(otros_materiales)
                
                # Dibujar los materiales utilizados en el PDF
                materiales_utilizados_chunks = [materiales_utilizados_completos[i:i+3] for i in range(0, len(materiales_utilizados_completos), 3)]
                for index, chunk in enumerate(materiales_utilizados_chunks):
                    c.drawString(100, 230 - index * 20, "Materiales Utilizados {}: {}".format(index+1, ", ".join(chunk)))
                
                c.drawString(100, 190, "Responsable Taller: {}".format(self.responsable_taller_var.get()))
                c.drawString(100, 170, "Responsable Recepcion: {}".format(self.responsable_recepcion_entry.get()))
                c.drawString(100, 150, "Correo: {}".format(self.correo_entry.get()))
                c.drawString(100, 125, "Firma Responsable Taller: ")
                c.drawString(390, 125, "Firma Responsable Equipo: ")
                
                c.drawImage("./img/logo1.png", letter[0] - 100, letter[1] - 70, width=100, height=50, mask='auto')
                
                c.drawImage("firma_Taller.png", 80, 50, width=120, height=70, mask='auto')
                c.drawImage("firma_Equipo.png", 390, 50, width=120, height=70, mask='auto')
                
                # Determinar la posición inicial en la segunda página
                y_start = 750  # Ajusta esta coordenada según tus necesidades

                # Dibujar encabezado de la segunda página
                c.showPage()  # Cambiar a la segunda página del PDF
                c.drawImage("./img/logo1.png", letter[0] - 100, letter[1] - 70, width=100, height=50, mask='auto')
                

                # Ajustar la posición vertical para los datos adicionales
                y = y_start - 50  # Ajusta esta cantidad según tus necesidades

                # Dibujar los datos adicionales en la segunda página
                c.drawString(100, y, "Motivo por el cual no fue reparado:")
                y -= 20  # Ajusta esta cantidad según tus necesidades

                # Dibujar los checkboxes de motivo por el cual no fue reparado
                for motivo_check, motivo_var in self.motivos_no_reparado:
                    if motivo_var.get():
                        
                        c.drawString(120, y, motivo_check)
                        
                        y -= 20  # Ajusta esta cantidad según tus necesidades
                c.drawString(120, y, otro_motivo)
                c.save()
                messagebox.showinfo("Ticket", "Se ha generado el ticket.")
                self.ticket_presionado = True
                # Habilitar el botón de guardar
                self.generar_ticket_button.config(state="disabled")
                self.correo_button.config(state="normal")
        
            # Crear ventana para ingresar motivo
            motivo_window = tk.Toplevel(self.root)
            motivo_window.title("Motivo por el cual no fue reparado")
            # Obtener la posición de la ventana principal
            x_main, y_main = self.root.winfo_x(), self.root.winfo_y()
        
            # Calcular la posición de la ventana emergente
            x_motivo = x_main + 1000  # Ajusta este valor según tu preferencia
            y_motivo = y_main + 100  # Ajusta este valor según tu preferencia
        
            # Establecer la posición de la ventana emergente
            motivo_window.geometry(f"+{x_motivo}+{y_motivo}")
        
            # Checkboxes 
            self.motivos_no_reparado = [
                ("Refacciones obsoletas", tk.BooleanVar()),
                ("Reparación no costeable", tk.BooleanVar()),
                ("Material no disponible en el taller", tk.BooleanVar()),
                ("Tarjeta dañada en su totalidad", tk.BooleanVar()),
                ("Espera de refacciones", tk.BooleanVar())
            ]
            for i, (motivo_check, motivo_var) in enumerate(self.motivos_no_reparado):
                tk.Checkbutton(motivo_window, text=motivo_check, variable=motivo_var).grid(row=i, column=0, sticky="w")
        
            otros_motivo_label = tk.Label(motivo_window, text="Otros:")
            otros_motivo_label.grid(row=len(self.motivos_no_reparado), column=0, sticky="w")
            otros_motivo_entry = tk.Entry(motivo_window)
            otros_motivo_entry.grid(row=len(self.motivos_no_reparado), column=1, sticky="w")
        
            guardar_button = tk.Button(motivo_window, text="Guardar", command=lambda: guardar_motivo(otros_motivo_entry))
            guardar_button.grid(row=len(self.motivos_no_reparado) + 1, column=1, columnspan=2)
    
            # Mostrar ventana de motivo
            motivo_window.mainloop()
    
        else:
            selected_item = self.equipo_combobox.get()
            # Generar el ticket
            numero = self.contador_label.cget("text")
            
            ticket_pdf_path =  f"./Archivos/PDFs/{numero}.pdf"
            c = canvas.Canvas(ticket_pdf_path, pagesize=letter)
            # Establecer la fuente en el lienzo
            
            c.drawString(200, 750, "Ticket Generado")
            c.drawString(100, 730, "Número de folio: {}".format(numero))
            c.drawString(100, 710, "Fecha de Recepcion: {}".format(self.fecha_recepcion_entry.get()))
            c.drawString(100, 690, "Fecha de Entrega: {}".format(self.fecha_entrega_entry.get()))
            c.drawString(100, 670, "Nombre del Responsable: {}".format(self.nombre_responsable_entry.get()))
            c.drawString(100, 650, "Teléfono del Responsable: {}".format(self.telefono_responsable_entry.get()))
            c.drawString(100, 630, "Área del Responsable: {}".format(self.area_responsable_entry.get()))
            c.drawString(100, 610, "Equipo: {}".format(selected_item))
            c.drawString(100, 590, "Otro Equipo: {}".format(self.otro_equipo_entry.get()))
            c.drawString(100, 570, "Modelo del Equipo: {}".format(self.modelo_equipo_entry.get()))
            c.drawString(100, 550, "Marca del Equipo: {}".format(self.marca_equipo_entry.get()))
            c.drawString(100, 530, "Número de Serie: {}".format(self.no_serie_entry.get()))
            c.drawString(100, 510, "Número de Inventario: {}".format(self.no_inventario_entry.get()))
            falta_equipo= self.faltante_equipo_entry.get()
            c.drawString(200, 490, "Estado Equipo")
            c.drawString(100, 470, "Enciende: {}".format("Sí" if self.estado_enciende.get() else "No"))
            c.drawString(100, 450, "Falta algun componente: {}".format((falta_equipo) if self.estado_componente.get() else "No"))
            c.drawString(100, 430, "Presenta Corrosion/Oxidacion: {}".format("Sí" if self.estado_corrosion.get() else "No"))
            c.drawString(100, 410, "Cable Alimentacion: {}".format("Sí" if self.estado_cable.get() else "No"))
            c.drawString(100, 390, "Daño Botones/Perillas: {}".format("Sí" if self.estado_dano_botones.get() else "No"))
            c.drawString(100, 370, "Daño Carcasa: {}".format("Sí" if self.estado_dano_carcasa.get() else "No"))
            c.drawString(100, 350, "Descripcion detallada: {}".format(self.descripcion_detallada_entry.get("1.0", tk.END).strip()))
            #c.drawString(200, 330, "Mantenimineto:")
            
            # Obtener los tipos de mantenimiento seleccionados
            tipos_mantenimiento = []
            if self.preventivo_var.get():
                tipos_mantenimiento.append("Preventivo")
            if self.correctivo_var.get():
                tipos_mantenimiento.append("Correctivo")
            if self.diagnostico_var.get():
                tipos_mantenimiento.append("Diagnóstico")
            if self.puesta_en_marcha_var.get():
                tipos_mantenimiento.append("Puesta en Marcha")
            if self.otro_var.get():
                tipos_mantenimiento.append("Otro")
            
            # Dibujar los tipos de mantenimiento en el PDF
            tipos_mantenimiento_chunks = [tipos_mantenimiento[i:i+3] for i in range(0, len(tipos_mantenimiento), 3)]
            for index, chunk in enumerate(tipos_mantenimiento_chunks):
                c.drawString(100, 330 - index * 20, "Tipos de Mantenimiento {}: {}".format(index+1, ", ".join(chunk)))
            
            c.drawString(100, 290, "Descripcion: {}".format(self.descripcion_entry.get("1.0", tk.END).strip()))
            c.drawString(100, 270, "¿Fue Reparado?: {}".format(self.reparado_var.get()))
            c.drawString(200, 250, "Materiales Utilizados")
            
            # Obtener los materiales utilizados seleccionados
            otros_materiales = self.otros_materiales_entry.get()
            materiales_utilizados = [nombre for nombre, var in self.materiales_utilizados if var.get()]
            materiales_utilizados_completos = materiales_utilizados.copy()  # Copia la lista original
            if otros_materiales:
                materiales_utilizados_completos.append(otros_materiales)
            
            # Dibujar los materiales utilizados en el PDF
            materiales_utilizados_chunks = [materiales_utilizados_completos[i:i+3] for i in range(0, len(materiales_utilizados_completos), 3)]
            for index, chunk in enumerate(materiales_utilizados_chunks):
                c.drawString(100, 230 - index * 20, "Materiales Utilizados {}: {}".format(index+1, ", ".join(chunk)))
            
            c.drawString(100, 190, "Responsable Taller: {}".format(self.responsable_taller_var.get()))
            c.drawString(100, 170, "Responsable Recepcion: {}".format(self.responsable_recepcion_entry.get()))
            c.drawString(100, 150, "Correo: {}".format(self.correo_entry.get()))
            c.drawString(100, 125, "Firma Responsable Taller: ")
            c.drawString(390, 125, "Firma Responsable Equipo: ")
            
            c.drawImage("./img/logo1.png", letter[0] - 100, letter[1] - 70, width=100, height=50, mask='auto')
            
            c.drawImage("firma_Taller.png", 80, 50, width=120, height=70, mask='auto')
            c.drawImage("firma_Equipo.png", 390, 50, width=120, height=70, mask='auto')
         
            c.save()
            messagebox.showinfo("Ticket", "Se ha generado el ticket.")
            self.ticket_presionado = True
            # Habilitar el botón de guardar
            self.generar_ticket_button.config(state="disabled")
            self.correo_button.config(state="normal")

            
    def guardar_datos(self, event=None):
        # Obtener el equipo seleccionado en el combobox
        descripcion_equipo = self.equipo_combobox.get()
    
        # Obtener el estado de los checkboxes
        estado_enciende = self.estado_enciende.get()
        estado_cable = self.estado_cable.get()
        estado_componente = self.estado_componente.get()
        estado_dano_botones = self.estado_dano_botones.get()
        estado_corrosion = self.estado_corrosion.get()
        estado_dano_carcasa = self.estado_dano_carcasa.get()
    
        # Si la descripción del equipo es "Otros", usar el valor de la entrada de texto
        if descripcion_equipo == "Otros":
            descripcion_equipo = self.otro_equipo_entry.get()
    
        datos = {
            "fecha_recepcion": self.fecha_recepcion_entry.get(),
            "fecha_entrega": self.fecha_entrega_entry.get(),
            "nombre_responsable": self.nombre_responsable_entry.get(),
            "telefono_responsable": self.telefono_responsable_entry.get(),
            "area_responsable": self.area_responsable_entry.get(),
            "correo": self.correo_entry.get(),
            "descripcion_equipo": descripcion_equipo,
            "otro_equipo": self.otro_equipo_entry.get(),
            "falta_equipo": self.faltante_equipo_entry.get(),
            "modelo_equipo": self.modelo_equipo_entry.get(),
            "marca_equipo": self.marca_equipo_entry.get(),
            "no_serie": self.no_serie_entry.get(),
            "no_inventario": self.no_inventario_entry.get(),
            "descripcion_detallada": self.descripcion_detallada_entry.get("1.0", tk.END).strip(),
            "estado_enciende": estado_enciende,
            "estado_cable": estado_cable,
            "estado_componente": estado_componente,
            "estado_dano_botones": estado_dano_botones,
            "estado_corrosion": estado_corrosion,
            "estado_dano_carcasa": estado_dano_carcasa,
            # Sección de Mantenimiento del Equipo
            "mantenimiento_preventivo": self.preventivo_var.get(),
            "mantenimiento_correctivo": self.correctivo_var.get(),
            "mantenimiento_diagnostico": self.diagnostico_var.get(),
            "mantenimiento_puesta_en_marcha": self.puesta_en_marcha_var.get(),
            "mantenimiento_otro": self.otro_var.get(),
            # Otros campos de información
            "descripcion": self.descripcion_entry.get("1.0", tk.END).strip(),
            "reparado": self.reparado_var.get(),
            # Materiales Utilizados
            "materiales_utilizados": {material: var.get() for material, var in self.materiales_utilizados},
            "otro_material_utilizado": self.otros_materiales_entry.get(),
            # Responsables
            "responsable_taller": self.responsable_taller_var.get(),
            "responsable_equipo_recepcion": self.responsable_recepcion_entry.get()
        }
    
        
        numero = self.contador_label.cget("text")
        with open(f"./Archivos/Progresos/{numero}.json", "w") as f:
            json.dump(datos, f)
            messagebox.showinfo("Guardar Progreso", "Progreso Guardado Correctamente")
            self.guardar_p_button.config(state="disabled")
            self.guardar_button.config(state="normal")

    
    def cargar_datos(self):
        try:
            folio = simpledialog.askstring("Cargar Datos", "Ingrese el folioa a cargar:")
            if folio is None:  # El usuario canceló la entrada
                return
    
            with open(f"./Archivos/Progresos/{folio}.json", "r") as f:
                datos = json.load(f)
                self.fecha_recepcion_entry.delete(0, tk.END)
                self.fecha_entrega_entry.delete(0, tk.END)
                self.no_serie_entry.delete(0, tk.END)
                self.no_inventario_entry.delete(0, tk.END)
                # Asignar los datos cargados a los campos correspondientes
                self.fecha_recepcion_entry.insert(0, datos.get("fecha_recepcion", ""))
                self.fecha_entrega_entry.insert(0, datos.get("fecha_entrega", ""))
                self.nombre_responsable_entry.insert(0, datos.get("nombre_responsable", ""))
                self.telefono_responsable_entry.insert(0, datos.get("telefono_responsable", ""))
                self.area_responsable_entry.insert(0, datos.get("area_responsable", ""))
                self.correo_entry.insert(0, datos.get("correo", ""))
                self.otro_equipo_entry.insert(0, datos.get("otro_equipo", ""))
    
                # Verificar si el campo "estado_equipo" es True y el campo "falta_equipo" es "Falta algun componente"
                estado_componente = datos.get("estado_componente", False)
    
                if estado_componente:
                    self.faltante_equipo_entry.config(state="normal")
                    self.faltante_equipo_entry.insert(0, datos.get("falta_equipo", ""))
                else:
                    self.faltante_equipo_entry.config(state="disabled")
    
                # Cargar los estados de los Checkbuttons
                self.modelo_equipo_entry.insert(0, datos.get("modelo_equipo", ""))
                self.marca_equipo_entry.insert(0, datos.get("marca_equipo", ""))
                self.no_serie_entry.insert(0, datos.get("no_serie", ""))
                self.no_inventario_entry.insert(0, datos.get("no_inventario", ""))
                self.descripcion_detallada_entry.insert("1.0", datos.get("descripcion_detallada", ""))
                # Si la descripción del equipo es "Otros", establecer el valor en el combobox y habilitar la entrada de texto
                descripcion_equipo = datos.get("descripcion_equipo", "")
                if descripcion_equipo == "Otros":
                    self.equipo_combobox.set("Otros")
                    self.otro_equipo_entry.config(state="normal")
                else:
                    self.equipo_combobox.set(descripcion_equipo)
                    self.otro_equipo_entry.config(state="disabled")
                # Cargar los estados de los Checkbuttons
                self.estado_enciende.set(datos.get("estado_enciende", False))
                self.estado_cable.set(datos.get("estado_cable", False))
                self.estado_componente.set(datos.get("estado_componente", False))
                self.estado_dano_botones.set(datos.get("estado_dano_botones", False))
                self.estado_corrosion.set(datos.get("estado_corrosion", False))
                self.estado_dano_carcasa.set(datos.get("estado_dano_carcasa", False))
                # Sección de Mantenimiento del Equipo
                self.preventivo_var.set(datos.get("mantenimiento_preventivo", False))
                self.correctivo_var.set(datos.get("mantenimiento_correctivo", False))
                self.diagnostico_var.set(datos.get("mantenimiento_diagnostico", False))
                self.puesta_en_marcha_var.set(datos.get("mantenimiento_puesta_en_marcha", False))
                self.otro_var.set(datos.get("mantenimiento_otro", False))
                # Otros campos de información
                self.descripcion_entry.delete("1.0", tk.END)
                self.descripcion_entry.insert("1.0", datos.get("descripcion", ""))
                self.reparado_var.set(datos.get("reparado", ""))
                # Materiales Utilizados
                for material, var in self.materiales_utilizados:
                    var.set(datos.get("materiales_utilizados", {}).get(material, False))
                self.otros_materiales_entry.delete(0, tk.END)
                self.otros_materiales_entry.insert(0, datos.get("otro_material_utilizado", ""))
                # Responsables
                self.responsable_taller_var.set(datos.get("responsable_taller", ""))
                self.responsable_recepcion_entry.delete(0, tk.END)
                self.responsable_recepcion_entry.insert(0, datos.get("responsable_equipo_recepcion", ""))
                messagebox.showinfo("Cargar Progreso", "Progreso Cargado Correctamente")
                self.contador_label.config(text=folio)
        except FileNotFoundError:
            messagebox.showinfo("Cargar Progreso", "Progreso NO Cargado Correctamente")
    

    def cargar_numero(self):
    # Intentar cargar el número guardado desde el archivo
    
        try:
            with open("./Folio/numero_guardado.txt", "r") as file:
                numero = int(file.read())
                
            return numero
        except FileNotFoundError:
            return 0
        except ValueError:
            messagebox.showwarning("Advertencia", "El archivo de número guardado está dañado.")
            return 0
        
    def guardar_numero(self, numero):
        # Guardar el número en el archivo
        with open("./Folio/numero_guardado.txt", "w") as file:
            file.write(str(numero))
        
    def toggle_recibir(self, selection):
        if selection == "Si":
            self.responsable_recepcion_entry.config(state="normal")
            self.responsable_recepcion_entry.delete(0, tk.END)
            self.responsable_recepcion_entry.insert(0, self.nombre_responsable_entry.get())
            self.responsable_recepcion_entry.config(state="disabled")
        elif selection == "No":
            self.responsable_recepcion_entry.config(state="normal")
            self.responsable_recepcion_entry.delete(0, tk.END)
            self.responsable_recepcion_entry.focus()

    def toggle_otros(self):
        # Habilitar o deshabilitar la caja de texto "Otros" según el estado del checkbox "Otros"
        if self.otro_var.get():
            self.otros_materiales_entry.config(state="normal")
        else:
            self.otros_materiales_entry.config(state="disabled")
            
    def toggle_falta(self):
        # Habilitar o deshabilitar la caja de texto "falta" según el estado del checkbox "Otros"
        if self.estado_componente.get():
            self.faltante_equipo_entry.config(state="normal")
        else:
            self.faltante_equipo_entry.config(state="disabled")

       
    
            
    def guardar_bitacora(self):
        # Obtener los datos de los campos de entrada
        numero = self.contador_label.cget("text")
        fecha_recepcion = self.fecha_recepcion_entry.get_date()
        fecha_entrega = self.fecha_entrega_entry.get_date()
        nombre_responsable = self.nombre_responsable_entry.get()
        telefono_responsable = self.telefono_responsable_entry.get()
        area_responsable = self.area_responsable_entry.get()
        descripcion_equipo = self.equipo_combobox.get()
        otro_equipo = self.otro_equipo_entry.get() if descripcion_equipo == "Otros" else ""
        falta_equipo = self.faltante_equipo_entry.get()
        modelo_equipo = self.modelo_equipo_entry.get()
        marca_equipo = self.marca_equipo_entry.get()
        no_serie = self.no_serie_entry.get()
        no_inventario = self.no_inventario_entry.get()
        descripcion_detallada = self.descripcion_detallada_entry.get("1.0", tk.END).strip()  # Obtener todo el texto
        descripcion = self.descripcion_entry.get("1.0", tk.END).strip()  # Obtener todo el texto
        otros_materiales = self.otros_materiales_entry.get()
        responsable_taller = self.responsable_taller_var.get()
        responsable_recepcion = self.responsable_recepcion_entry.get()
        reparado = self.reparado_var.get()
    
        # Obtener los tipos de mantenimiento seleccionados
        tipos_mantenimiento = []
        if self.preventivo_var.get():
            tipos_mantenimiento.append("Preventivo")
        if self.correctivo_var.get():
            tipos_mantenimiento.append("Correctivo")
        if self.diagnostico_var.get():
            tipos_mantenimiento.append("Diagnóstico")
        if self.puesta_en_marcha_var.get():
            tipos_mantenimiento.append("Puesta en Marcha")
        if self.otro_var.get():
            tipos_mantenimiento.append("Otro")
    
        # Obtener los materiales utilizados seleccionados
        materiales_utilizados = [nombre for nombre, var in self.materiales_utilizados if var.get()]
        materiales_utilizados_completos = materiales_utilizados.copy()  # Copia la lista original
        if otros_materiales:
            materiales_utilizados_completos.append(otros_materiales)
    
        # Abrir el archivo de Excel existente o crear uno nuevo
        try:
            libro_excel = load_workbook("bitacora_mantenimiento.xlsx")
            hoja_activa = libro_excel.active
        except FileNotFoundError:
            libro_excel = Workbook()
            hoja_activa = libro_excel.active
            # Si el archivo no existe, crear encabezados
            encabezados = ["No.", "Fecha de Recepción", "Fecha de Entrega", "Nombre del Responsable", "Teléfono del Responsable",
                           "Área del Responsable", "Equipo", "Modelo", "Marca", "No.Serie", "No.Inventario",
                           "Enciende", "Cable de Alimentación",
                           "Falta algún componente", "Daño en botones/perillas",
                           "Presenta corrosión/oxidación", "Daño en Carcasa",
                           "Descripción Detallada", "Tipos de Mantenimiento", "Descripción", "¿Fue Reparado?",
                           "Lista de Materiales Utilizados", "Responsable del Taller", "Responsable equipo Recepcion", "Firma"]
            hoja_activa.append(encabezados)
    
        # Encontrar la primera fila vacía
        fila_vacia = hoja_activa.max_row + 1
    
        # Escribir los datos en la fila vacía
        hoja_activa.cell(row=fila_vacia, column=1).value = numero
        hoja_activa.cell(row=fila_vacia, column=2).value = fecha_recepcion
        hoja_activa.cell(row=fila_vacia, column=3).value = fecha_entrega
        hoja_activa.cell(row=fila_vacia, column=4).value = nombre_responsable
        hoja_activa.cell(row=fila_vacia, column=5).value = telefono_responsable
        hoja_activa.cell(row=fila_vacia, column=6).value = area_responsable
        hoja_activa.cell(row=fila_vacia, column=7).value = otro_equipo if descripcion_equipo == "Otros" else descripcion_equipo
    
        hoja_activa.cell(row=fila_vacia, column=8).value = modelo_equipo
        hoja_activa.cell(row=fila_vacia, column=9).value = marca_equipo
        hoja_activa.cell(row=fila_vacia, column=10).value = no_serie
        hoja_activa.cell(row=fila_vacia, column=11).value = no_inventario
        hoja_activa.cell(row=fila_vacia, column=12).value = "Sí" if self.estado_enciende.get() else "No"
        hoja_activa.cell(row=fila_vacia, column=13).value = "Sí" if self.estado_cable.get() else "No"
        hoja_activa.cell(row=fila_vacia, column=14).value = "Sí, {}".format(falta_equipo) if self.estado_componente.get() else "No"
        hoja_activa.cell(row=fila_vacia, column=15).value = "Sí" if self.estado_dano_botones.get() else "No"
        hoja_activa.cell(row=fila_vacia, column=16).value = "Sí" if self.estado_corrosion.get() else "No"
        hoja_activa.cell(row=fila_vacia, column=17).value = "Sí" if self.estado_dano_carcasa.get() else "No"
        hoja_activa.cell(row=fila_vacia, column=18).value = descripcion_detallada
        hoja_activa.cell(row=fila_vacia, column=19).value = ", ".join(tipos_mantenimiento)
        hoja_activa.cell(row=fila_vacia, column=20).value = descripcion
        hoja_activa.cell(row=fila_vacia, column=21).value = reparado
        hoja_activa.cell(row=fila_vacia, column=22).value = ", ".join(materiales_utilizados_completos)
        hoja_activa.cell(row=fila_vacia, column=23).value = responsable_taller
        hoja_activa.cell(row=fila_vacia, column=24).value = responsable_recepcion
        # Guardar cambios en el archivo
        libro_excel.save("bitacora_mantenimiento.xlsx")

        # Mostrar mensaje de éxito
        messagebox.showinfo("Bitácora Guardada", "Los datos de la bitácora se han guardado correctamente.")
        
        # Limpiar campos después de guardar
        
        self.limpiar_campos()
        #self.numero_guardado += 1
        #self.guardar_numero(self.numero_guardado)
        self.contador_label.config(text=f"{self.numero_guardado}") 
        #self.generar_codigo_barras(str(self.numero_guardado))
        self.estadisticas_button.config(state="normal")
    
    def limpiar_campos(self):
        # Limpiar todos los campos de entrada
        self.fecha_recepcion_entry.delete(0, tk.END)
        self.fecha_entrega_entry.delete(0, tk.END)
        self.nombre_responsable_entry.delete(0, tk.END)
        self.telefono_responsable_entry.delete(0, tk.END)
        self.area_responsable_entry.delete(0, tk.END)
        self.correo_entry.delete(0, tk.END)
        self.modelo_equipo_entry.delete(0, tk.END)
        self.marca_equipo_entry.delete(0, tk.END)
        self.no_serie_entry.delete(0, tk.END)
        
        self.no_inventario_entry.delete(0, tk.END)
        
        self.descripcion_detallada_entry.delete("1.0", tk.END)
        self.descripcion_entry.delete("1.0", tk.END)
        self.otros_materiales_entry.delete(0, tk.END)
        self.otros_materiales_entry.config(state="disabled")
        
        self.otro_equipo_entry.delete(0, tk.END)
        self.otro_equipo_entry.config(state="disabled")
        self.responsable_recepcion_entry.delete(0, tk.END)
        self.faltante_equipo_entry.delete(0, tk.END)
        self.faltante_equipo_entry.config(state="disabled")       
    
        # Limpiar estados de los checkbuttons
        self.estado_enciende.set(False)
        self.estado_cable.set(False)
        self.estado_componente.set(False)
        self.estado_dano_botones.set(False)
        self.estado_corrosion.set(False)
        self.estado_dano_carcasa.set(False)
        self.preventivo_var.set(False)
        self.correctivo_var.set(False)
        self.diagnostico_var.set(False)
        self.puesta_en_marcha_var.set(False)
        self.otro_var.set(False)
        
        self.reparado_var.set("")
    
        # Limpiar materiales utilizados
        for _, var in self.materiales_utilizados:
            var.set(False)
    
        # Reiniciar la opción seleccionada en el menú desplegable
        self.responsable_taller_var.set("")
        self.equipo_combobox.set("")  # Reiniciar el ComboBox
        self.reparado_var.set("")
    
        # Desactivar el botón de guardar después de limpiar los campos
        self.guardar_button.config(state="disabled")
        self.generar_ticket_button.config(state="disabled")
        self.no_serie_var = tk.StringVar(value="N/A")
        self.no_inventario_var = tk.StringVar(value="N/A")

    def abrir_excel(self):
        # Abrir el archivo de Excel con la aplicación predeterminada
        try:
            os.startfile("bitacora_mantenimiento.xlsx")
        except FileNotFoundError:
            messagebox.showerror("Error", "El archivo de Excel no se ha encontrado.")
            
        pass

# Inicializar la aplicación Tkinter
root = tk.Tk()
root.geometry("900x750")
app = BitacoraMantenimiento(root)


# Ejecutar la aplicación
root.mainloop()
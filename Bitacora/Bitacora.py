from PIL import Image, ImageTk

import tkinter as tk
from tkinter import Scrollbar
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from barcode import Code39
from barcode.writer import ImageWriter
from tkcalendar import DateEntry  # Importamos el widget de calendario
import os


        
class AutocompleteEntry(tk.Entry):
    def __init__(self, parent, lista, *args, **kwargs):
        self.lista = lista
        self.var = tk.StringVar()
        tk.Entry.__init__(self, parent, textvariable=self.var, *args, **kwargs)
        self.var.trace('w', self.autocomplete)
        self.bind('<Return>', self.select)
        self.lb_up = False

    def autocomplete(self, *args):
        palabras = self.lista
        texto = self.var.get()
        if texto == '':
            self.lb.destroy()
            self.lb_up = False
        else:
            sugerencias = []
            for palabra in palabras:
                if palabra.lower().startswith(texto.lower()):
                    sugerencias.append(palabra)
            if sugerencias:
                if not self.lb_up:
                    self.lb = tk.Listbox(width=self.winfo_width())
                    self.lb.bind('<Double-Button-1>', self.select)
                    self.lb.bind('<Return>', self.select)
                    self.lb.grid(row=self.grid_info()['row'] + 1, column=self.grid_info()['column'], sticky='ew')
                    self.lb_up = True
                self.lb.delete(0, tk.END)
                for sugerencia in sugerencias:
                    self.lb.insert(tk.END, sugerencia)
            else:
                if self.lb_up:
                    self.lb.destroy()
                    self.lb_up = False

    def select(self, *args):
        if self.lb_up:
            self.var.set(self.lb.get(tk.ACTIVE))
            self.lb.destroy()
            self.lb_up = False
            self.icursor(tk.END)    

class BitacoraMantenimiento:
    def __init__(self, root):   
        
        self.root = root
        self.root.title("Bitácora de Mantenimiento")
        self.root.iconbitmap("logo4.ico")
        # Crear una variable de control para el contador

        
        # Crear un lienzo para colocar el frame con la barra de desplazamiento
        canvas = tk.Canvas(root)
        canvas.pack(side='left', fill='both', expand=True)
        
        # Crear un frame dentro del lienzo
        self.frame = tk.Frame(canvas)
        self.frame.pack(padx=10, pady=10)
        
        # Añadir la barra de desplazamiento vertical
        scrollbar = tk.Scrollbar(root, orient='vertical', command=canvas.yview)
        scrollbar.pack(side='right', fill='y')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Configurar el lienzo para usar la barra de desplazamiento
        canvas.create_window((0, 0), window=self.frame, anchor='nw')
        
        # Vincular la barra de desplazamiento con el movimiento del ratón
        self.frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        # Enlazar el desplazamiento del mouse al lienzo
        canvas.bind_all("<MouseWheel>", lambda event: canvas.yview_scroll(int(-1*(event.delta/120)), "units"))

        
        # Agregar contenido al frame
        self.add_content()
        
 

    def add_content(self):
        
        #self.contador = tk.IntVar(value=2)
        # Intentar cargar el número guardado desde el archivo
        self.numero_guardado = self.cargar_numero()
        # Crear una etiqueta para mostrar el contador
        self.contador_label = tk.Label(root, text=f"{self.numero_guardado}")
        self.contador_label.place(x=850, y=10)  # Colocar la etiqueta en la esquina superior derecha

                  # Cargar las imágenes
        self.logo1_image = Image.open("logo1.png")
        self.logo1_image = self.logo1_image.resize((180, 100), )  # Redimensionar la imagen
        self.logo1_photo = ImageTk.PhotoImage(self.logo1_image)
        
        self.logo2_image = Image.open("logo2.png")
        self.logo2_image = self.logo2_image.resize((180, 100), )  # Redimensionar la imagen
        self.logo2_photo = ImageTk.PhotoImage(self.logo2_image)
        
        # Etiqueta para la primera imagen
        self.logo1_label = tk.Label(self.frame, image=self.logo1_photo)
        self.logo1_label.grid(row=0, column=0, padx=10, pady=10)
        
        # Etiqueta para la segunda imagen
        self.logo2_label = tk.Label(self.frame, image=self.logo2_photo)
        self.logo2_label.grid(row=0, column=3, padx=10, pady=10)

        
        # Título para la sección de Información del Equipo
        titulo_informacion_equipo = tk.Label(self.frame, text="Bitacora de Mantenimiento", font=("Arial", 14, "bold"), pady=10)
        titulo_informacion_equipo.grid(row=0, column=0, columnspan=4)
        titulo_informacion_equipo.config(justify="center")
        
        # Campos de entrada
        tk.Label(self.frame, text="Fecha de Recepción:").grid(row=1, column=0, sticky="e")
        self.fecha_recepcion_entry = DateEntry(self.frame, width=12, background='darkblue', foreground='white', borderwidth=2)
        self.fecha_recepcion_entry.grid(row=1, column=1)
        
        tk.Label(self.frame, text="Fecha de Entrega:").grid(row=1, column=2, sticky="e")
        self.fecha_entrega_entry = DateEntry(self.frame, width=12, background='darkblue', foreground='white', borderwidth=2)
        self.fecha_entrega_entry.grid(row=1, column=3)
        
        # Título 
        titulo_datos_equipo = tk.Label(self.frame, text="Datos del Responsable del Equipo", font=("Helvetica", 12, "bold"), pady=10)
        titulo_datos_equipo.grid(row=2, column=0, columnspan=4)
        titulo_datos_equipo.config(justify="center")
        
        tk.Label(self.frame, text="Nombre:").grid(row=3, column=0, sticky="e")
        self.nombre_responsable_entry = tk.Entry(self.frame)
        self.nombre_responsable_entry.grid(row=3, column=1)
        
        tk.Label(self.frame, text="Teléfono:").grid(row=4, column=0, sticky="e")
        self.telefono_responsable_entry = tk.Entry(self.frame)
        self.telefono_responsable_entry.grid(row=4, column=1)
        
        tk.Label(self.frame, text="Área:").grid(row=4, column=2, sticky="e")
        self.area_responsable_entry = tk.Entry(self.frame)
        self.area_responsable_entry.grid(row=4, column=3)
        
        # Título 
        titulo_descripcion_equipo = tk.Label(self.frame, text="Descripcion de Equipo", font=("Helvetica", 12, "bold"), pady=10)
        titulo_descripcion_equipo.grid(row=6, column=0, columnspan=4)
        titulo_descripcion_equipo.config(justify="center")
        tk.Label(self.frame, text="Equipo:").grid(row=7, column=0, sticky="e")

        self.equipo_var = tk.StringVar()
        equipo_options = ["Autoclave", "Balanza digital", "Balanza Granataria", "Basculas", "Bocina", "Bomba de agua", "Bomba de vacío", 
                          "Campana", "Cámaras", "Cargador", "Centrifugadora", "Cortadora", "Compresora", "Computadora", "Control Remoto", "Cronometro", 
                          "Dispensador de agua", "Diseño 3D", "Eliminador", "Esmeril", "Espectrofotómetro", "Estufa", "Fuentes de poder", "Generador de funciones", 
                          "Horno", "Impresión 3D", "Impresora", "Incubadora", "Lámpara", "Licuadora", "Lavadoras", "Mano factura de piezas", "Manta de calentamiento", 
                          "Microscopio", "Microscopio Estereoscópico", "Monitor", "Motor", "Mufla", "Multímetro", "Multi-contactos", "No break", "Osciloscopio","Pantalla", 
                          "Parilla con agitación", "Parrilla de calentamiento", "Parrilla de calentamiento y agitación", "Pipetas", "Proyector", "Pulidora", "Purificador", 
                          "Radio", "Refrigerador", "Regulador de Voltaje", "Sensores", "Termómetro digital", "Taladro", "Torneo y fresado", "Tornos","Torniquete","UPS","Otros" ]

        scrollbar = Scrollbar(self.frame, orient='vertical')
        equipo_listbox = tk.Listbox(self.frame, yscrollcommand=scrollbar.set)
        scrollbar.config(command=equipo_listbox.yview)
        scrollbar.grid(row=7, column=2, sticky='ns')
        equipo_listbox.grid(row=7, column=1, sticky='ew')
        
        for option in equipo_options:
            equipo_listbox.insert(tk.END, option)
            
       
        equipo_listbox.bind("<<ListboxSelect>>", lambda event, lb=equipo_listbox: self.toggle_equipo(event, lb))

       
        tk.Label(self.frame, text="Otros:").grid(row=7, column=2, sticky="e")
        self.otro_equipo_entry = tk.Entry(self.frame, state="disabled")
        self.otro_equipo_entry.grid(row=7, column=3)

                
        
        tk.Label(self.frame, text="Modelo:").grid(row=8, column=0, sticky="e")
        self.modelo_equipo_entry = tk.Entry(self.frame)
        self.modelo_equipo_entry.grid(row=8, column=1)
        
        tk.Label(self.frame, text="Marca:").grid(row=8, column=2, sticky="e")
        self.marca_equipo_entry = tk.Entry(self.frame)
        self.marca_equipo_entry.grid(row=8, column=3)
        
        tk.Label(self.frame, text="No. Serie").grid(row=10, column=0, sticky="e")
        self.no_serie_var = tk.StringVar(value="N/A")
        self.no_serie_entry = tk.Entry(self.frame, textvariable=self.no_serie_var)
        self.no_serie_entry.grid(row=10, column=1)
        
        
        
        tk.Label(self.frame, text="No. Inventario").grid(row=10, column=2, sticky="e")
        self.no_inventario_var = tk.StringVar(value="N/A")
        self.no_inventario_entry = tk.Entry(self.frame, textvariable=self.no_inventario_var)
        self.no_inventario_entry.grid(row=10, column=3)
        
        # Título 
        titulo_estado_equipo = tk.Label(self.frame, text="Estado del Equipo", font=("Helvetica", 12, "bold"), pady=10)
        titulo_estado_equipo.grid(row=11, column=0, columnspan=4)
        titulo_estado_equipo.config(justify="center")
        
        # Sección de Estado del Equipo
        tk.Label(self.frame, text="Estado del Equipo:").grid(row=12, column=0, sticky="e")
        self.estado_enciende = tk.BooleanVar()
        self.enciende_checkbutton = tk.Checkbutton(self.frame, text="Enciende", variable=self.estado_enciende)
        self.enciende_checkbutton.grid(row=12, column=1, sticky="w")
        
        self.estado_cable = tk.BooleanVar()
        self.cable_checkbutton = tk.Checkbutton(self.frame, text="Cable de Alimentación", variable=self.estado_cable)
        self.cable_checkbutton.grid(row=12, column=2, sticky="w")
        
        self.estado_componente = tk.BooleanVar()
        self.componente_checkbutton = tk.Checkbutton(self.frame, text="Falta algún componente", variable=self.estado_componente,command=self.toggle_falta)
        self.componente_checkbutton.grid(row=13, column=1, sticky="w")
        
        
        self.faltante_equipo_entry = tk.Entry(self.frame, state="disabled")
        self.faltante_equipo_entry.grid(row=13, column=3)
        
        
        self.estado_dano_botones = tk.BooleanVar()
        self.dano_botones_checkbutton = tk.Checkbutton(self.frame, text="Daño en botones/perillas", variable=self.estado_dano_botones)
        self.dano_botones_checkbutton.grid(row=13, column=2, sticky="w")
        
        self.estado_corrosion = tk.BooleanVar()
        self.corrosion_checkbutton = tk.Checkbutton(self.frame, text="Presenta corrosión/oxidación", variable=self.estado_corrosion)
        self.corrosion_checkbutton.grid(row=14, column=1, sticky="w")
        
        self.estado_dano_carcasa = tk.BooleanVar()
        self.dano_carcasa_checkbutton = tk.Checkbutton(self.frame, text="Daño en Carcasa", variable=self.estado_dano_carcasa)
        self.dano_carcasa_checkbutton.grid(row=14, column=2, sticky="w")
        
        # Título 
        titulo_estado_equipo = tk.Label(self.frame, text="Descripción Detallada del Equipo", font=("Helvetica", 12, "bold"), pady=10)
        titulo_estado_equipo.grid(row=15, column=0, columnspan=4)
        titulo_estado_equipo.config(justify="center")
        
        # Otros campos de información
        tk.Label(self.frame, text="Descripción Detallada del Equipo:").grid(row=16, column=0, sticky="ne")
        self.descripcion_detallada_entry = tk.Text(self.frame, height=5, width=80, wrap=tk.WORD)
        self.descripcion_detallada_entry.grid(row=16, column=1, columnspan=3, sticky="w")
        
        # Título 
        titulo_estado_equipo = tk.Label(self.frame, text="Mantenimineto", font=("Helvetica", 12, "bold"), pady=10)
        titulo_estado_equipo.grid(row=17, column=0, columnspan=4)
        titulo_estado_equipo.config(justify="center")
        
        # Sección de Mantenimiento del Equipo
        tk.Label(self.frame, text="Mantenimiento del Equipo:").grid(row=18, column=0, sticky="e")
        self.preventivo_var = tk.BooleanVar()
        self.preventivo_checkbutton = tk.Checkbutton(self.frame, text="Preventivo", variable=self.preventivo_var)
        self.preventivo_checkbutton.grid(row=18, column=1, sticky="w")
        
        self.correctivo_var = tk.BooleanVar()
        self.correctivo_checkbutton = tk.Checkbutton(self.frame, text="Correctivo", variable=self.correctivo_var)
        self.correctivo_checkbutton.grid(row=18, column=2, sticky="w")
        
        self.diagnostico_var = tk.BooleanVar()
        self.diagnostico_checkbutton = tk.Checkbutton(self.frame, text="Diagnóstico", variable=self.diagnostico_var)
        self.diagnostico_checkbutton.grid(row=19, column=1, sticky="w")
        
        self.puesta_en_marcha_var = tk.BooleanVar()
        self.puesta_en_marcha_checkbutton = tk.Checkbutton(self.frame, text="Puesta en Marcha", variable=self.puesta_en_marcha_var)
        self.puesta_en_marcha_checkbutton.grid(row=19, column=2, sticky="w")
        
        self.otro_var = tk.BooleanVar()
        self.otro_checkbutton = tk.Checkbutton(self.frame, text="Otro", variable=self.otro_var)
        self.otro_checkbutton.grid(row=19, column=3, sticky="w")
        
        # Otros campos de información
        tk.Label(self.frame, text="Descripción:").grid(row=20, column=0, sticky="ne")
        self.descripcion_entry = tk.Text(self.frame, height=5, width=80, wrap=tk.WORD)
        self.descripcion_entry.grid(row=20, column=1, columnspan=3, sticky="w")
        
        tk.Label(self.frame, text="¿Fue reparado?").grid(row=21, column=0, sticky="e")
        self.reparado_var = tk.BooleanVar()
        self.reparado_checkbutton = tk.Checkbutton(self.frame, variable=self.reparado_var)
        self.reparado_checkbutton.grid(row=21, column=1, sticky="w")
        
        tk.Label(self.frame, text="SI/NO").grid(row=21, column=1, sticky="e")
        
        # Título 
        titulo_estado_equipo = tk.Label(self.frame, text="Lista de Materiales Utilizados", font=("Helvetica", 12, "bold"), pady=10)
        titulo_estado_equipo.grid(row=22, column=0, columnspan=4)
        titulo_estado_equipo.config(justify="center")
        
        # Checkboxes para Materiales Utilizados
        tk.Label(self.frame, text="Materiales Utilizados:").grid(row=23, column=0, sticky="e")
        self.materiales_utilizados = [
            ("Aceite lubricante multiusos", tk.BooleanVar()),
            ("Alcohol Isopropílico", tk.BooleanVar()),
            ("Soldadura (Estaño)", tk.BooleanVar()),
            ("Aislantes", tk.BooleanVar()),
            ("Cable de Conexión", tk.BooleanVar()),
            ("Conectores y/o terminales", tk.BooleanVar()),
            ("Potenciómetros", tk.BooleanVar()),
            ("Cables de Alimentación", tk.BooleanVar()),
            ("Dispositivos electrónicos de potencia", tk.BooleanVar()),
            ("Dispositivos de sujeción", tk.BooleanVar()),
            ("Fusibles", tk.BooleanVar()),
            ("Liquido Limpiador Multiusos", tk.BooleanVar())
        ]
        for i, (nombre, var) in enumerate(self.materiales_utilizados):
            if i < 6:
                column = 1
            else:
                column = 2
                i -= 6
            tk.Checkbutton(self.frame, text=nombre, variable=var).grid(row=23+i, column=column, sticky="w")
            
        self.otro_var = tk.BooleanVar()
        self.otro_checkbox = tk.Checkbutton(self.frame, text="Otros", variable=self.otro_var,command=self.toggle_otros)
        self.otro_checkbox.grid(row=30, column=2, sticky="w")   
        

        
        # Caja de texto para Otros Materiales Utilizados
        tk.Label(self.frame, text="Otros:").grid(row=30, column=0, sticky="e")
        self.otros_materiales_entry = tk.Entry(self.frame, state="disabled")
        self.otros_materiales_entry.grid(row=30, column=1, columnspan=2, sticky="w")
        

        
        # Título 
        titulo_estado_equipo = tk.Label(self.frame, text="Responsables", font=("Helvetica", 12, "bold"), pady=10)
        titulo_estado_equipo.grid(row=31, column=0, columnspan=4)
        titulo_estado_equipo.config(justify="center")
        
        
                # Responsable de Taller
        tk.Label(self.frame, text="Responsable de Taller:").grid(row=32, column=0, sticky="e")
        self.responsable_taller_var = tk.StringVar()
        responsable_taller_options = ["Juan Daniel Ramírez Zamora", "Jose Manuel Fernandez Ramírez", "Otro"]
        self.responsable_taller_dropdown = tk.OptionMenu(self.frame, self.responsable_taller_var, *responsable_taller_options)
        self.responsable_taller_dropdown.grid(row=32, column=1, columnspan=2, sticky="w")
        
        # Cajas de texto para Responsable equipo Recepción y Firma de conformidad
        tk.Label(self.frame, text="Responsable equipo Recepción:").grid(row=32, column=2, sticky="e")
        self.responsable_recepcion_entry = tk.Entry(self.frame)
        self.responsable_recepcion_entry.grid(row=32, column=3, columnspan=2, sticky="w")
        

        
        # Botón para guardar
        
        self.guardar_button = tk.Button(self.frame, text="Guardar", command=lambda: self.guardar_bitacora(lb=equipo_listbox))

        self.guardar_button.grid(row=34, column=0, columnspan=3, pady=10)
        
        # Botón para abrir el archivo de Excel
        self.abrir_button = tk.Button(self.frame, text="Abrir Excel", command=self.abrir_excel)
        self.abrir_button.grid(row=34, column=2, columnspan=3, pady=10)
        
        self.generar_ticket_button = tk.Button(self.frame,text="Generar Ticket", command=lambda: self.generar_ticket_y_codigo_barras(lb=equipo_listbox))
        self.generar_ticket_button.grid(row=34, column=3, columnspan=3, pady=10)


               


        
        # Desactivar el botón de guardar al principio
        self.guardar_button.config(state="disabled")
                  
    def toggle_equipo(self, event, lb):
        selected_index = lb.curselection()  # Obtener el índice del elemento seleccionado
        if selected_index:  # Verificar si se ha seleccionado algún elemento
            selected_item = lb.get(selected_index[0])  # Obtener el elemento seleccionado
            if selected_item == "Otros":
                self.otro_equipo_entry.config(state="normal")  # Activar la caja de texto "otro_equipo"
            else:
                self.otro_equipo_entry.config(state="disabled")  # Desactivar la caja de texto "otro_equipo"

  
    def generar_ticket_y_codigo_barras(self,event=None, lb=None):
        selected_index = lb.curselection()
        if selected_index:  # Verificar si se ha seleccionado algún elemento
            descripcion_equipo = lb.get(selected_index[0])
        else:
            descripcion_equipo = "" 
            
        # Generar el ticket
        numero = self.numero_guardado
        ticket_pdf_path =  f"./Archivos/PDFs/{numero}.pdf"
        c = canvas.Canvas(ticket_pdf_path, pagesize=letter)
        c.drawString(200, 750, "Ticket Generado")
        c.drawString(100, 730, "Número de ticket: {}".format(self.numero_guardado))
        c.drawString(100, 710, "Fecha de Recepcion: {}".format(self.fecha_recepcion_entry.get()))
        c.drawString(100, 690, "Fecha de Entrega: {}".format(self.fecha_entrega_entry.get()))
        c.drawString(100, 670, "Nombre del Responsable: {}".format(self.nombre_responsable_entry.get()))
        c.drawString(100, 650, "Teléfono del Responsable: {}".format(self.telefono_responsable_entry.get()))
        c.drawString(100, 630, "Área del Responsable: {}".format(self.area_responsable_entry.get()))
        c.drawString(100, 610, "Equipo: {}".format(descripcion_equipo))
        c.drawString(100, 590, "Otro Equipo: {}".format(self.otro_equipo_entry.get()))
        c.drawString(100, 570, "Modelo del Equipo: {}".format(self.modelo_equipo_entry.get()))
        c.drawString(100, 550, "Marca del Equipo: {}".format(self.marca_equipo_entry.get()))
        c.drawString(100, 530, "Número de Serie: {}".format(self.no_serie_entry.get()))
        c.drawString(100, 510, "Número de Inventario: {}".format(self.no_inventario_entry.get()))
        falta_equipo= self.faltante_equipo_entry.get()
        c.drawString(200, 490, "Estado Equipo:")
        c.drawString(100, 470, "Enciende: {}".format("Sí" if self.estado_enciende.get() else "No"))
        c.drawString(100, 450, "Falta algun componente: {}".format((falta_equipo) if self.estado_componente.get() else "No"))
        c.drawString(100, 430, "Presenta Corrosion/Oxidacion: {}".format("Sí" if self.estado_corrosion.get() else "No"))
        c.drawString(100, 410, "Cable Alimentacion: {}".format("Sí" if self.estado_cable.get() else "No"))
        c.drawString(100, 390, "Daño Botones/Perillas: {}".format("Sí" if self.estado_dano_botones.get() else "No"))
        c.drawString(100, 370, "Daño Carcasa: {}".format("Sí" if self.estado_dano_carcasa.get() else "No"))
        c.drawString(100, 350, "Descripcion detallada: {}".format(self.descripcion_detallada_entry.get("1.0", tk.END).strip()))
        c.drawString(200, 330, "Mantenimineto:")
        # Obtener los tipos de mantenimiento seleccionados
        tipos_mantenimiento = []
        if self.preventivo_var.get():
            tipos_mantenimiento.append("Preventivo")
        if self.correctivo_var.get():
            tipos_mantenimiento.append("Correctivo")
        if self.diagnostico_var.get():
            tipos_mantenimiento.append("Diagnóstico")
        if self.puesta_en_marcha_var.get():
            tipos_mantenimiento.append("Puesta en Marcha")
        if self.otro_var.get():
            tipos_mantenimiento.append("Otro")
        c.drawString(100, 330, "Tipos: {}".format(", ".join(tipos_mantenimiento)))
        c.drawString(100, 310, "Descripcion: {}".format(self.descripcion_entry.get("1.0", tk.END).strip()))
        c.drawString(100, 290, "¿Fue Reparado?: {}".format("Sí" if self.reparado_var.get() else "No"))
        c.drawString(200, 270, "Materiales Utilizados")
        # Obtener los materiales utilizados seleccionados
        otros_materiales = self.otros_materiales_entry.get()
        materiales_utilizados = [nombre for nombre, var in self.materiales_utilizados if var.get()]
        materiales_utilizados_completos = materiales_utilizados.copy()  # Copia la lista original
        if otros_materiales:
            materiales_utilizados_completos.append(otros_materiales)
        c.drawString(100, 250, "Tipos: {}".format(", ".join(materiales_utilizados_completos)))
        c.drawString(100, 230, "Responsable Taller: {}".format(self.responsable_taller_var.get()))
        c.drawString(100, 210, "Responsable Recepcion: {}".format(self.responsable_recepcion_entry.get()))
       
        # Generar el código de barras
        codigo_barras = Code39(str(self.numero_guardado), writer=ImageWriter())
        codigo_barras_file = "codigo_barras.png"

        codigo_barras.save(codigo_barras_file)
        
        # Insertar el código de barras en el ticket
        c.drawImage(codigo_barras_file, 100, 100, width=200, height=50)  # Ajusta la posición según sea necesario
        
        c.save()

        # Mostrar un mensaje de confirmación
        messagebox.showinfo("Ticket y Código de Barras Generados", "Se ha generado el ticket y el código de barras.")
        
    def cargar_numero(self):
    # Intentar cargar el número guardado desde el archivo
        try:
            with open("numero_guardado.txt", "r") as file:
                numero = int(file.read())
            return numero
        except FileNotFoundError:
            return 0
        except ValueError:
            messagebox.showwarning("Advertencia", "El archivo de número guardado está dañado.")
            return 0
        
    def guardar_numero(self, numero):
        # Guardar el número en el archivo
        with open("numero_guardado.txt", "w") as file:
            file.write(str(numero))
        
    
    def toggle_otros(self):
        # Habilitar o deshabilitar la caja de texto "Otros" según el estado del checkbox "Otros"
        if self.otro_var.get():
            self.otros_materiales_entry.config(state="normal")
        else:
            self.otros_materiales_entry.config(state="disabled")
            
    def toggle_falta(self):
        # Habilitar o deshabilitar la caja de texto "falta" según el estado del checkbox "Otros"
        if self.estado_componente.get():
            self.faltante_equipo_entry.config(state="normal")
        else:
            self.faltante_equipo_entry.config(state="disabled")
            
   

    def check_campos(self):
        # Verificar si hay algo escrito en todas las cajas de texto
        campos = [
            self.fecha_recepcion_entry.get(),
            self.fecha_entrega_entry.get(),
            self.nombre_responsable_entry.get(),
            self.telefono_responsable_entry.get(),
            self.area_responsable_entry.get(),
            self.modelo_equipo_entry.get(),
            self.marca_equipo_entry.get(),
            #self.no_serie_entry.get(),
            #self.no_inventario_entry.get(),
            self.descripcion_detallada_entry.get("1.0", tk.END).strip(),
            self.descripcion_entry.get("1.0", tk.END).strip(),
            #self.otros_materiales_entry.get(),
            self.responsable_taller_var.get(),
            self.responsable_recepcion_entry.get(),
            
        ]
        # Comprobar si alguna caja de texto está vacía
        if all(campos):
            self.guardar_button.config(state="normal")
        else:
            self.guardar_button.config(state="disabled")
            
    
            


    def guardar_bitacora(self,event=None, lb=None):
        # Obtener los datos de los campos de entrada
        numero = self.numero_guardado
        fecha_recepcion = self.fecha_recepcion_entry.get_date()
        fecha_entrega = self.fecha_entrega_entry.get_date()
        nombre_responsable = self.nombre_responsable_entry.get()
        telefono_responsable = self.telefono_responsable_entry.get()
        area_responsable = self.area_responsable_entry.get()
        selected_index = lb.curselection()  # Obtener el índice del elemento seleccionado
        if selected_index:  # Verificar si se ha seleccionado algún elemento
            descripcion_equipo = lb.get(selected_index[0])
        else:
            descripcion_equipo = "" 
        otro_equipo= self.otro_equipo_entry.get()
        falta_equipo= self.faltante_equipo_entry.get()
        modelo_equipo = self.modelo_equipo_entry.get()
        marca_equipo = self.marca_equipo_entry.get()
        no_serie = self.no_serie_entry.get()
        no_inventario = self.no_inventario_entry.get()
        descripcion_detallada = self.descripcion_detallada_entry.get("1.0", tk.END).strip()  # Obtener todo el texto
        descripcion = self.descripcion_entry.get("1.0", tk.END).strip()  # Obtener todo el texto
        otros_materiales = self.otros_materiales_entry.get()
        responsable_taller = self.responsable_taller_var.get()
        responsable_recepcion = self.responsable_recepcion_entry.get()
        
    

        # Obtener los tipos de mantenimiento seleccionados
        tipos_mantenimiento = []
        if self.preventivo_var.get():
            tipos_mantenimiento.append("Preventivo")
        if self.correctivo_var.get():
            tipos_mantenimiento.append("Correctivo")
        if self.diagnostico_var.get():
            tipos_mantenimiento.append("Diagnóstico")
        if self.puesta_en_marcha_var.get():
            tipos_mantenimiento.append("Puesta en Marcha")
        if self.otro_var.get():
            tipos_mantenimiento.append("Otro")

        # Obtener los materiales utilizados seleccionados
        materiales_utilizados = [nombre for nombre, var in self.materiales_utilizados if var.get()]
        
        
        materiales_utilizados_completos = materiales_utilizados.copy()  # Copia la lista original
        if otros_materiales:
            materiales_utilizados_completos.append(otros_materiales)

        # Abrir el archivo de Excel existente o crear uno nuevo
        try:
            libro_excel = load_workbook("bitacora_mantenimiento.xlsx")
            hoja_activa = libro_excel.active
        except FileNotFoundError:
            libro_excel = Workbook()
            hoja_activa = libro_excel.active
            # Si el archivo no existe, crear encabezados
            encabezados = ["No.","Fecha de Recepción", "Fecha de Entrega", "Nombre del Responsable", "Teléfono del Responsable",
                           "Área del Responsable", "Equipo", "Modelo", "Marca", "No.Serie", "No.Inventario",
                           "Enciende", "Cable de Alimentación",
                           "Falta algún componente", "Daño en botones/perillas",
                           "Presenta corrosión/oxidación", "Daño en Carcasa",
                           "Descripción Detallada", "Tipos de Mantenimiento", "Descripción", "¿Fue Reparado?",
                           "Lista de Materiales Utilizados", "Responsable del Taller", "Responsable equipo Recepcion","Firma"]
            hoja_activa.append(encabezados)
            

        # Encontrar la primera fila vacía
        fila_vacia = hoja_activa.max_row + 1

        # Escribir los datos en la fila vacía
        hoja_activa.cell(row=fila_vacia, column=1).value = numero
        hoja_activa.cell(row=fila_vacia, column=2).value = fecha_recepcion
        hoja_activa.cell(row=fila_vacia, column=3).value = fecha_entrega
        hoja_activa.cell(row=fila_vacia, column=4).value = nombre_responsable
        hoja_activa.cell(row=fila_vacia, column=5).value = telefono_responsable
        hoja_activa.cell(row=fila_vacia, column=6).value = area_responsable
        hoja_activa.cell(row=fila_vacia, column=7).value = otro_equipo if descripcion_equipo == "Otros" else descripcion_equipo
        
        
        hoja_activa.cell(row=fila_vacia, column=8).value = modelo_equipo
        hoja_activa.cell(row=fila_vacia, column=9).value = marca_equipo
        hoja_activa.cell(row=fila_vacia, column=10).value = no_serie
        hoja_activa.cell(row=fila_vacia, column=11).value = no_inventario
        hoja_activa.cell(row=fila_vacia, column=12).value = "Sí" if self.estado_enciende.get() else "No"
        hoja_activa.cell(row=fila_vacia, column=13).value = "Sí" if self.estado_cable.get() else "No"
        hoja_activa.cell(row=fila_vacia, column=14).value = "Sí, {}".format(falta_equipo) if self.estado_componente.get() else "No"
        hoja_activa.cell(row=fila_vacia, column=15).value = "Sí" if self.estado_dano_botones.get() else "No"
        hoja_activa.cell(row=fila_vacia, column=16).value = "Sí" if self.estado_corrosion.get() else "No"
        hoja_activa.cell(row=fila_vacia, column=17).value = "Sí" if self.estado_dano_carcasa.get() else "No"
        hoja_activa.cell(row=fila_vacia, column=18).value = descripcion_detallada
        hoja_activa.cell(row=fila_vacia, column=19).value = ", ".join(tipos_mantenimiento)
        hoja_activa.cell(row=fila_vacia, column=20).value = descripcion
        hoja_activa.cell(row=fila_vacia, column=21).value = "Sí" if self.reparado_var.get() else "No"
        hoja_activa.cell(row=fila_vacia, column=22).value = ", ".join(materiales_utilizados_completos)
        hoja_activa.cell(row=fila_vacia, column=23).value = responsable_taller
        hoja_activa.cell(row=fila_vacia, column=24).value = responsable_recepcion
        
        if not self.reparado_var.get():  # Si no fue reparado
            motivo_window = tk.Toplevel(self.root)
            motivo_window.title("Motivo por el cual no fue reparado")
            # Obtener la posición de la ventana principal
            x_main, y_main = self.root.winfo_x(), self.root.winfo_y()
            
            # Calcular la posición de la ventana emergente
            x_motivo = x_main + 1000  # Ajusta este valor según tu preferencia
            y_motivo = y_main + 100  # Ajusta este valor según tu preferencia
            
            # Establecer la posición de la ventana emergente
            motivo_window.geometry(f"+{x_motivo}+{y_motivo}")
    
            def guardar_motivo():
                motivo = ""
                for motivo_check, motivo_var in self.motivos_no_reparado:
                    if motivo_var.get():
                        motivo += motivo_check + ", "
                otro_motivo = otros_motivo_entry.get()
                if otro_motivo:
                    motivo += otro_motivo + ", "
    
                    # Obtener el número de contador
                    numero = self.numero_guardado
                
                    # Encontrar la última fila en la segunda hoja
                    hoja_motivos = libro_excel["Motivos_No_Reparado1"]
                    last_row = len(hoja_motivos['A'])
                
                    # Guardar el numero y el motivo en la segunda hoja
                    hoja_motivos.cell(row=last_row + 1, column=1, value=numero)
                    hoja_motivos.cell(row=last_row + 1, column=2, value=motivo)
                
                    # Guardar cambios en el archivo
                    libro_excel.save("bitacora_mantenimiento.xlsx")
                
                    messagebox.showinfo("Motivo Guardado", f"El motivo por el cual no fue reparado es: {motivo}")
                
                    motivo_window.destroy()
    
            # Checkboxes 
            self.motivos_no_reparado = [
                ("Refacciones obsoletas", tk.BooleanVar()),
                ("Reparación no costeable", tk.BooleanVar()),
                ("Material no disponible en el taller", tk.BooleanVar()),
                ("Tarjeta dañada en su totalidad", tk.BooleanVar()),
                ("Espera de refacciones", tk.BooleanVar())
            ]
            for i, (motivo_check, motivo_var) in enumerate(self.motivos_no_reparado):
                tk.Checkbutton(motivo_window, text=motivo_check, variable=motivo_var).grid(row=i, column=0, sticky="w")
    
            otros_motivo_label = tk.Label(motivo_window, text="Otros:")
            otros_motivo_label.grid(row=len(self.motivos_no_reparado), column=0, sticky="w")
            otros_motivo_entry = tk.Entry(motivo_window)
            otros_motivo_entry.grid(row=len(self.motivos_no_reparado), column=1, sticky="w")
    
            guardar_button = tk.Button(motivo_window, text="Guardar", command=guardar_motivo)
            guardar_button.grid(row=len(self.motivos_no_reparado) + 1, column=1, columnspan=2)

        # Guardar cambios en el archivo
        libro_excel.save("bitacora_mantenimiento.xlsx")

        # Mostrar mensaje de éxito
        messagebox.showinfo("Bitácora Guardada", "Los datos de la bitácora se han guardado correctamente.")
        
        # Limpiar campos después de guardar
        
        self.limpiar_campos()
        self.numero_guardado += 1
        self.guardar_numero(self.numero_guardado)
        
       
    
    def limpiar_campos(self):
        # Limpiar todos los campos de entrada
        self.fecha_recepcion_entry.delete(0, tk.END)
        self.fecha_entrega_entry.delete(0, tk.END)
        self.nombre_responsable_entry.delete(0, tk.END)
        self.telefono_responsable_entry.delete(0, tk.END)
        self.area_responsable_entry.delete(0, tk.END)
        
        self.modelo_equipo_entry.delete(0, tk.END)
        self.marca_equipo_entry.delete(0, tk.END)
        self.no_serie_entry.delete(0, tk.END)
        
        self.no_inventario_entry.delete(0, tk.END)
        
        self.descripcion_detallada_entry.delete("1.0", tk.END)
        self.descripcion_entry.delete("1.0", tk.END)
        self.otros_materiales_entry.delete(0, tk.END)
        self.otro_equipo_entry.delete(0, tk.END)
        self.responsable_recepcion_entry.delete(0, tk.END)
        self.faltante_equipo_entry.delete(0, tk.END)
        
    
        # Limpiar estados de los checkbuttons
        self.estado_enciende.set(False)
        self.estado_cable.set(False)
        self.estado_componente.set(False)
        self.estado_dano_botones.set(False)
        self.estado_corrosion.set(False)
        self.estado_dano_carcasa.set(False)
        self.preventivo_var.set(False)
        self.correctivo_var.set(False)
        self.diagnostico_var.set(False)
        self.puesta_en_marcha_var.set(False)
        self.otro_var.set(False)
        self.reparado_var.set(False)
    
        # Limpiar materiales utilizados
        for _, var in self.materiales_utilizados:
            var.set(False)
    
        # Reiniciar la opción seleccionada en el menú desplegable
        self.responsable_taller_var.set("")
        self.equipo_var.set("")
    
        # Desactivar el botón de guardar después de limpiar los campos
        self.guardar_button.config(state="disabled")
        self.no_serie_var = tk.StringVar(value="N/A")
        self.no_inventario_var = tk.StringVar(value="N/A")



    def abrir_excel(self):
        # Abrir el archivo de Excel con la aplicación predeterminada
        try:
            os.startfile("bitacora_mantenimiento.xlsx")
        except FileNotFoundError:
            messagebox.showerror("Error", "El archivo de Excel no se ha encontrado.")
            
        pass

# Inicializar la aplicación Tkinter
root = tk.Tk()
root.geometry("900x550")
app = BitacoraMantenimiento(root)

# Verificar los campos en cada cambio
root.bind("<Key>", lambda e: app.check_campos())

# Ejecutar la aplicación
root.mainloop()